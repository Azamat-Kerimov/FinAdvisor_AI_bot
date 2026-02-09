# FastAPI сервер для Telegram Web App
# v_01.28.26 - Рефакторинг: полная бизнес-логика, проверка подписки, AI
from __future__ import annotations

from fastapi import FastAPI, HTTPException, Depends, Header, Request, Query, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
import traceback
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List
import tempfile
import asyncpg
import os
from dotenv import load_dotenv
import hmac
import hashlib
import json
import uuid
import base64
import asyncio
import httpx
from datetime import datetime, timedelta
from decimal import Decimal

_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
load_dotenv(_env_path)

APP_ENV = (os.getenv("APP_ENV") or "").strip().lower()

# Настройка логирования
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

app = FastAPI()

# Статика фронта (React/Vite build) — frontend/dist
_frontend_dist = os.path.join(os.path.dirname(__file__), "frontend", "dist")
_frontend_ready = False
if os.path.isdir(_frontend_dist):
    _assets = os.path.join(_frontend_dist, "assets")
    index_path = os.path.join(_frontend_dist, "index.html")
    if os.path.isdir(_assets) and os.path.isfile(index_path):
        try:
            app.mount("/assets", StaticFiles(directory=_assets), name="assets")
            _frontend_ready = True
        except Exception as e:
            logging.warning("Could not mount frontend assets: %s", e)
    _examples = os.path.join(_frontend_dist, "examples")
    if os.path.isdir(_examples):
        try:
            app.mount("/examples", StaticFiles(directory=_examples), name="examples")
        except Exception as e:
            logging.warning("Could not mount frontend examples: %s", e)
if not _frontend_ready:
    logging.warning(
        "Frontend not found at %s — run: ./scripts/build_frontend.sh",
        _frontend_dist,
    )

# При необработанном исключении: в тесте возвращаем детали в теле 500 (для отладки автотестов)
@app.exception_handler(Exception)
async def unhandled_exception_handler(request, exc):
    logging.exception("Unhandled exception: %s", exc)
    if APP_ENV == "test":
        return JSONResponse(
            status_code=500,
            content={"detail": str(exc), "traceback": traceback.format_exc()},
        )
    return JSONResponse(status_code=500, content={"detail": "Internal Server Error"})

# CORS для Telegram Web App
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database connection
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
BOT_TOKEN = os.getenv("BOT_TOKEN")

# GigaChat credentials
G_CLIENT_ID = os.getenv("GIGACHAT_CLIENT_ID")
G_CLIENT_SECRET = os.getenv("GIGACHAT_CLIENT_SECRET")
G_SCOPE = os.getenv("GIGACHAT_SCOPE")
G_AUTH_URL = os.getenv("GIGACHAT_AUTH_URL")
G_API_URL = os.getenv("GIGACHAT_API_URL")
GIGACHAT_MODEL = os.getenv("GIGACHAT_MODEL", "GigaChat:2.0.28.2")

db_pool: Optional[asyncpg.Pool] = None


def _json_serializable(val):
    """Привести значение из asyncpg (Decimal, date) к типу, сериализуемому в JSON."""
    if isinstance(val, Decimal):
        return float(val)
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def _row_to_dict(r: asyncpg.Record) -> dict:
    """Преобразовать запись asyncpg в dict с JSON-сериализуемыми значениями."""
    return {k: _json_serializable(r[k]) for k in r.keys()}


# Кэш для bot_module, чтобы избежать повторного импорта
_bot_module_cache = None

async def get_db():
    global db_pool
    if db_pool is None:
        db_pool = await asyncpg.create_pool(
            user=DB_USER, password=DB_PASSWORD, database=DB_NAME,
            host=DB_HOST, port=DB_PORT, min_size=1, max_size=6
        )
    return db_pool

# Telegram Web App validation
def validate_telegram_webapp(init_data: str) -> dict:
    """Проверка подписи Telegram Web App"""
    import urllib.parse
    
    try:
        # URL декодируем init_data (на случай, если он пришел закодированным)
        init_data_decoded = urllib.parse.unquote(init_data)
        
        # Парсим initData
        params = {}
        for item in init_data_decoded.split('&'):
            if '=' in item:
                key, value = item.split('=', 1)
                # Декодируем значение (может быть закодировано несколько раз)
                params[key] = urllib.parse.unquote(value)
        
        # Проверяем hash
        hash_value = params.pop('hash', '')
        if not hash_value:
            raise HTTPException(status_code=401, detail="Missing hash in initData")
        
        # Создаем строку для проверки (важно: сортировка по ключам)
        data_check_string = '\n'.join(f"{k}={v}" for k, v in sorted(params.items()))
        
        # Создаем секретный ключ
        secret_key = hmac.new(
            "WebAppData".encode(), 
            BOT_TOKEN.encode(), 
            hashlib.sha256
        ).digest()
        
        # Вычисляем хеш
        calculated_hash = hmac.new(
            secret_key,
            data_check_string.encode(),
            hashlib.sha256
        ).hexdigest()
        
        # Сравниваем хеши
        if calculated_hash != hash_value:
            import logging
            logging.error(f"Hash mismatch. Expected: {hash_value}, Got: {calculated_hash}")
            logging.error(f"Data check string: {data_check_string[:100]}...")
            raise HTTPException(status_code=401, detail="Invalid hash")
        
        # Парсим user
        user_str = params.get('user', '')
        if not user_str:
            raise HTTPException(status_code=401, detail="Missing user in initData")
        
        user = json.loads(user_str) if user_str else {}
        
        return user
    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=401, detail=f"Invalid user JSON: {str(e)}")
    except Exception as e:
        import logging
        logging.error(f"Validation error: {e}")
        import traceback
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=401, detail=f"Validation error: {str(e)}")


def _username_from_telegram_user(user: dict) -> Optional[str]:
    """Из объекта user в initData достать username (snake_case или camelCase)."""
    v = user.get('username') or user.get('userName')
    return (v or '').strip() or None


def _display_name_from_telegram_user(user: dict) -> Optional[str]:
    """Имя для отображения: username или first_name + last_name (snake/camelCase)."""
    u = _username_from_telegram_user(user)
    if u:
        return u
    first = (user.get('first_name') or user.get('firstName') or '').strip()
    last = (user.get('last_name') or user.get('lastName') or '').strip()
    return (' '.join((first, last)).strip() or None) if (first or last) else None


def _is_test_user_request(request: Request) -> bool:
    """Проверка: запрос с тестовым user_id (без Telegram)."""
    return APP_ENV == "test" and bool(request.headers.get("x-test-user-id"))


async def get_user_id(request: Request) -> int:
    """Получить user_id из Telegram Web App или в тесте из заголовка X-Test-User-Id"""
    # В тестовой среде разрешаем заголовок X-Test-User-Id (без Telegram)
    if APP_ENV == "test":
        test_user_id = request.headers.get("x-test-user-id")
        if test_user_id:
            try:
                uid = int(test_user_id)
                if uid > 0:
                    return uid
            except ValueError:
                pass

    # Пробуем получить init-data из заголовков (nginx может передавать как init-data или init_data)
    init_data = request.headers.get("init-data") or request.headers.get("init_data")

    if not init_data:
        logging.warning("Missing init-data header in request")
        raise HTTPException(status_code=401, detail="Missing initData. Откройте приложение через Telegram.")

    try:
        user = validate_telegram_webapp(init_data)
        tg_id = user.get('id')
        
        if not tg_id:
            raise HTTPException(status_code=401, detail="Invalid user")
        
        # Получаем или создаем пользователя (username или first_name + last_name)
        display_name = _display_name_from_telegram_user(user)
        db = await get_db()
        async with db.acquire() as conn:
            row = await conn.fetchrow("SELECT id, username FROM users WHERE tg_id=$1", tg_id)
            if not row:
                await conn.execute(
                    "INSERT INTO users (tg_id, username, created_at) VALUES ($1, $2, NOW())",
                    tg_id, display_name
                )
                row = await conn.fetchrow("SELECT id FROM users WHERE tg_id=$1", tg_id)
            else:
                # Обновить username, если в БД пусто, а в initData есть имя
                if display_name and (not row['username'] or row['username'] != display_name):
                    await conn.execute(
                        "UPDATE users SET username=$1 WHERE tg_id=$2",
                        display_name, tg_id
                    )
            return row['id']
    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in get_user_id: {e}")
        raise HTTPException(status_code=401, detail=f"Authentication error: {str(e)}")


async def check_premium(user_id: int) -> bool:
    """Проверить активна ли подписка"""
    db = await get_db()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT premium_until FROM users WHERE id=$1", user_id
        )
        if not row or not row['premium_until']:
            return False
        return row['premium_until'] > datetime.now()


async def require_premium(request: Request, user_id: int = Depends(get_user_id)):
    """Dependency для проверки подписки - возвращает user_id если подписка активна. В тесте с X-Test-User-Id подписку не проверяем."""
    if _is_test_user_request(request):
        return user_id
    if not await check_premium(user_id):
        raise HTTPException(
            status_code=403,
            detail="PREMIUM_REQUIRED"
        )
    return user_id


# Для гибрида: приложение доступно и без подписки (консультации лимитированы 1/мес). Используем get_user_id везде.
# require_premium оставлен для обратной совместиости; эндпоинты переведены на get_user_id.

# Pydantic models
class TransactionCreate(BaseModel):
    amount: float
    category_id: Optional[int] = None  # предпочтительно
    category: Optional[str] = None     # имя категории (если category_id не передан)
    description: Optional[str] = None

class TransactionUpdate(BaseModel):
    amount: Optional[float] = None
    category_id: Optional[int] = None
    category: Optional[str] = None
    description: Optional[str] = None

class GoalCreate(BaseModel):
    title: str
    target: float
    description: Optional[str] = None

class AssetCreate(BaseModel):
    title: str
    type: str
    amount: float

class AssetUpdate(BaseModel):
    title: Optional[str] = None
    type: Optional[str] = None
    amount: Optional[float] = None

class LiabilityCreate(BaseModel):
    title: str
    type: str
    amount: float
    monthly_payment: float

class LiabilityUpdate(BaseModel):
    title: Optional[str] = None
    type: Optional[str] = None
    amount: Optional[float] = None
    monthly_payment: Optional[float] = None


class TransactionImportItem(BaseModel):
    """Один элемент импорта (category_id из справочника/маппинга)"""
    date: str  # YYYY-MM-DD
    amount: float
    category_id: int
    description: Optional[str] = None


class ImportApplyRequest(BaseModel):
    mode: str  # "add" | "replace"
    transactions: List[TransactionImportItem]


class ConsultationMessageRequest(BaseModel):
    message: str


class BudgetCreate(BaseModel):
    category: str
    monthly_limit: float


class ProfileUpdate(BaseModel):
    gender: Optional[str] = None
    birth_date: Optional[str] = None  # YYYY-MM-DD
    marital_status: Optional[str] = None
    children_count: Optional[int] = None
    city: Optional[str] = None


class LogActionRequest(BaseModel):
    """Тело запроса для логирования действия пользователя."""
    action: str
    details: Optional[dict] = None


# --- Профиль пользователя ---

@app.get("/api/profile")
async def get_profile(user_id: int = Depends(get_user_id)):
    """Получить профиль пользователя (пол, дата рождения, семья, город)."""
    db = await get_db()
    try:
        async with db.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT gender, birth_date, marital_status, children_count, city FROM users WHERE id=$1",
                user_id
            )
            if not row:
                raise HTTPException(status_code=404, detail="User not found")
            return {
                "gender": row["gender"],
                "birth_date": row["birth_date"].isoformat() if row.get("birth_date") else None,
                "marital_status": row["marital_status"],
                "children_count": row["children_count"],
                "city": row["city"],
            }
    except asyncpg.UndefinedColumnError:
        return {"gender": None, "birth_date": None, "marital_status": None, "children_count": None, "city": None}


@app.put("/api/profile")
async def update_profile(body: ProfileUpdate, user_id: int = Depends(get_user_id)):
    """Обновить профиль пользователя."""
    birth_date_parsed = None
    if body.birth_date is not None:
        if body.birth_date == "":
            birth_date_parsed = None
        else:
            try:
                from datetime import date
                parts = body.birth_date.strip().split("-")
                if len(parts) == 3:
                    birth_date_parsed = date(int(parts[0]), int(parts[1]), int(parts[2]))
            except (ValueError, IndexError):
                raise HTTPException(status_code=400, detail="Invalid birth_date format (use YYYY-MM-DD)")
    try:
        db = await get_db()
        async with db.acquire() as conn:
            await conn.execute(
                """
                UPDATE users SET
                    gender = COALESCE($1, gender),
                    birth_date = COALESCE($2, birth_date),
                    marital_status = COALESCE($3, marital_status),
                    children_count = COALESCE($4, children_count),
                    city = COALESCE($5, city)
                WHERE id = $6
                """,
                body.gender if body.gender else None,
                birth_date_parsed,
                body.marital_status if body.marital_status else None,
                body.children_count,
                body.city if body.city else None,
                user_id,
            )
    except asyncpg.UndefinedColumnError:
        raise HTTPException(status_code=400, detail="Profile columns not found. Run scripts/migrate_users_profile.sql")
    return {"status": "ok"}


@app.delete("/api/me/data")
async def delete_all_user_data(user_id: int = Depends(get_user_id)):
    """Удалить все данные пользователя (транзакции, цели, активы, долги, кэш ИИ и т.д.), кроме записи в users."""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            await conn.execute("DELETE FROM budgets WHERE user_id = $1", user_id)
        except asyncpg.UndefinedTableError:
            pass
        await conn.execute("DELETE FROM ai_cache WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM ai_context WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM goals WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM transactions WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM asset_values WHERE asset_id IN (SELECT id FROM assets WHERE user_id = $1)", user_id)
        await conn.execute("DELETE FROM assets WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM liability_values WHERE liability_id IN (SELECT id FROM liabilities WHERE user_id = $1)", user_id)
        await conn.execute("DELETE FROM liabilities WHERE user_id = $1", user_id)
        try:
            await conn.execute("DELETE FROM user_consultation_actions WHERE user_id = $1", user_id)
        except asyncpg.UndefinedTableError:
            pass
        try:
            await conn.execute("DELETE FROM user_focus_goal WHERE user_id = $1", user_id)
        except asyncpg.UndefinedTableError:
            pass
        try:
            await conn.execute("DELETE FROM user_actions WHERE user_id = $1", user_id)
        except asyncpg.UndefinedTableError:
            pass
    return {"status": "ok", "message": "Все данные удалены. Профиль сохранён."}


# --- Логирование действий пользователей ---

@app.post("/api/log-action")
async def log_action(body: LogActionRequest, user_id: int = Depends(get_user_id)):
    """Записать действие пользователя в БД (экран, шаринг, и т.д.)."""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            await conn.execute(
                "INSERT INTO user_actions (user_id, action, details, created_at) VALUES ($1, $2, $3, NOW())",
                user_id,
                (body.action or "").strip() or "unknown",
                body.details,
            )
        except asyncpg.UndefinedTableError:
            logging.warning("user_actions table not found; run migrate_user_actions.sql")
    return {"status": "ok"}


# --- Справочник категорий (БД) ---

@app.get("/api/categories")
async def get_categories_list(user_id: int = Depends(get_user_id)):
    """Список категорий приложения (id, name, type) из БД."""
    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, name, type FROM categories ORDER BY type, name"
        )
        return [{"id": r["id"], "name": r["name"], "type": r["type"]} for r in rows]


async def _get_category_id_by_name(conn, name: str) -> int | None:
    """Вернуть id категории по имени или None."""
    if not name or not str(name).strip():
        return None
    row = await conn.fetchrow(
        "SELECT id FROM categories WHERE name = $1",
        str(name).strip()
    )
    return row["id"] if row else None


async def _resolve_bank_category_to_id(
    conn, bank_category: str, amount: float, is_expense_row: bool | None = None
) -> int:
    """
    Резолв категории банка (Сбер/Т-Банк) в category_id по маппингу (bank_category + bank_category_type).
    is_expense_row: True = расход, False = доход, None = по знаку amount.
    """
    key = (bank_category or "").strip().lower()
    if not key:
        fallback = "Прочие расходы" if (is_expense_row if is_expense_row is not None else amount < 0) else "Прочие доходы"
        row = await conn.fetchrow("SELECT id FROM categories WHERE name = $1", fallback)
        return row["id"] if row else 1
    cat_type = "Расход" if (is_expense_row if is_expense_row is not None else amount < 0) else "Доход"
    row = await conn.fetchrow(
        """SELECT category_id FROM category_mapping
           WHERE LOWER(TRIM(bank_category)) = $1 AND bank_category_type = $2""",
        key, cat_type
    )
    if row:
        return row["category_id"]
    fallback_name = "Прочие расходы" if cat_type == "Расход" else "Прочие доходы"
    fallback_row = await conn.fetchrow("SELECT id FROM categories WHERE name = $1", fallback_name)
    fallback_id = fallback_row["id"] if fallback_row else 1
    try:
        await conn.execute(
            """INSERT INTO category_mapping (bank_category, category_id, bank_category_type)
               VALUES ($1, $2, $3)""",
            key, fallback_id, cat_type
        )
    except asyncpg.UniqueViolationError:
        pass
    return fallback_id


# API Endpoints

# Информация о среде только в тесте (какая БД, чтобы не перепутать с продом)
@app.get("/api/env-info")
async def get_env_info():
    """В тесте возвращает environment и данные подключения к БД. В проде — 404."""
    if APP_ENV != "test":
        raise HTTPException(status_code=404, detail="Not available")
    return {
        "environment": "test",
        "db_name": DB_NAME or "",
        "db_host": DB_HOST or "",
    }


# Auth endpoint (без проверки подписки)
@app.post("/api/auth/telegram")
async def auth_telegram(request: Request):
    """Авторизация через Telegram Web App initData"""
    init_data = request.headers.get("init-data") or request.headers.get("init_data")
    
    if not init_data:
        raise HTTPException(status_code=401, detail="Missing initData")
    
    try:
        user = validate_telegram_webapp(init_data)
        tg_id = user.get('id')
        
        if not tg_id:
            raise HTTPException(status_code=401, detail="Invalid user")
        
        # Получаем или создаем пользователя с 2 бесплатными месяцами (username или first_name + last_name)
        display_name = _display_name_from_telegram_user(user)
        db = await get_db()
        async with db.acquire() as conn:
            row = await conn.fetchrow("SELECT id, premium_until, username FROM users WHERE tg_id=$1", tg_id)
            if not row:
                # Новый пользователь - даем 2 бесплатных месяца
                free_months_until = datetime.now() + timedelta(days=60)
                await conn.execute(
                    "INSERT INTO users (tg_id, username, created_at, premium_until) VALUES ($1, $2, NOW(), $3)",
                    tg_id, display_name, free_months_until
                )
                row = await conn.fetchrow("SELECT id, premium_until FROM users WHERE tg_id=$1", tg_id)
            else:
                # Обновить username, если в БД пусто, а в initData есть имя
                if display_name and (not row['username'] or row['username'] != display_name):
                    await conn.execute(
                        "UPDATE users SET username=$1 WHERE tg_id=$2",
                        display_name, tg_id
                    )
            
            premium_until = row['premium_until']
            premium_active = premium_until and premium_until > datetime.now()
            
            return {
                "user_id": row['id'],
                "premium_until": premium_until.isoformat() if premium_until else None,
                "premium_active": premium_active
            }
    except HTTPException:
        raise
    except Exception as e:
        import logging
        logging.error(f"Error in auth_telegram: {e}")
        raise HTTPException(status_code=401, detail=f"Authentication error: {str(e)}")


@app.get("/yandex_eb705230b9f963ce.html", response_class=HTMLResponse)
async def yandex_verification():
    """Файл верификации домена для Яндекс.Вебмастер (должен лежать в frontend/dist после сборки)."""
    path = os.path.join(_frontend_dist, "yandex_eb705230b9f963ce.html")
    if os.path.isfile(path):
        with open(path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    raise HTTPException(status_code=404, detail="File not found. Rebuild frontend so public/yandex_*.html is in dist.")


@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Главная страница — React SPA из frontend/dist или заглушка"""
    index_path = os.path.join(_frontend_dist, "index.html")
    if os.path.isfile(index_path):
        try:
            with open(index_path, "r", encoding="utf-8") as f:
                content = f.read()
            headers = {
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            }
            return HTMLResponse(content=content, headers=headers)
        except Exception:
            pass
    return HTMLResponse(
        content="<h1>FinAdvisor API</h1><p>Фронт не собран. На сервере из корня проекта выполните: <code>./scripts/build_frontend.sh</code>, затем перезапустите API.</p>",
        status_code=200
    )


# Статистика
@app.get("/api/stats")
async def get_stats(
    month: Optional[int] = None,
    year: Optional[int] = None,
    user_id: int = Depends(get_user_id)
):
    """Получить статистику за выбранный месяц (по умолчанию — предыдущий)."""
    from datetime import datetime, date
    now = datetime.now()
    if month is None or year is None:
        # Предыдущий месяц
        first_this = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        prev = first_this - timedelta(days=1)
        year, month = prev.year, prev.month
    start = date(year, month, 1)
    end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)

    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT t.amount, c.name AS category, t.created_at
            FROM transactions t
            JOIN categories c ON c.id = t.category_id
            WHERE t.user_id=$1 AND t.created_at >= $2 AND t.created_at < $3
            ORDER BY t.created_at ASC
            """,
            user_id, start, end
        )
        
        income_by_cat = {}
        expense_by_cat = {}
        
        for r in rows:
            amount = float(r["amount"])
            cat = r["category"] or "—"
            if amount >= 0:
                income_by_cat[cat] = income_by_cat.get(cat, 0) + amount
            else:
                expense_by_cat[cat] = expense_by_cat.get(cat, 0) + (-amount)
        
        total_income = sum(income_by_cat.values())
        total_expense = sum(expense_by_cat.values())
        
        # Ценность 4: рекомендуемый резервный фонд (3 мес. расходов)
        reserve_recommended = round(total_expense * 3, 0) if total_expense else 0
        
        # Ценность 5: короткий инсайт по топу расходов
        top_expense = sorted(expense_by_cat.items(), key=lambda x: -x[1])[:3]
        total_exp = total_expense or 1
        insight_parts = [f"{cat}: {int(amt):,} ₽ ({int(100 * amt / total_exp)}%)".replace(",", " ") for cat, amt in top_expense]
        insight = "Топ расходов за месяц: " + ", ".join(insight_parts) if insight_parts else "Пока нет расходов за месяц."

        # Явно приводим к типам, сериализуемым в JSON (избегаем Decimal и т.п.)
        return {
            "month": month,
            "year": year,
            "total_income": float(total_income),
            "total_expense": float(total_expense),
            "income_by_category": {k: float(v) for k, v in income_by_cat.items()},
            "expense_by_category": {k: float(v) for k, v in expense_by_cat.items()},
            "reserve_recommended": int(reserve_recommended),
            "insight": insight,
        }


@app.get("/api/stats/monthly")
async def get_stats_monthly(user_id: int = Depends(get_user_id)):
    """Доходы, расходы и разница по месяцам за последние 12 месяцев."""
    from datetime import date
    now = datetime.now()
    result = []
    month_names_ru = (
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    )
    db = await get_db()
    async with db.acquire() as conn:
        for i in range(11, -1, -1):
            m = now.month - 1 - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            start = date(y, m, 1)
            end = date(y, m + 1, 1) if m < 12 else date(y + 1, 1, 1)
            rows = await conn.fetch(
                """
                SELECT amount FROM transactions t
                WHERE t.user_id=$1 AND t.created_at >= $2 AND t.created_at < $3
                """,
                user_id, start, end
            )
            income = sum(float(r["amount"]) for r in rows if float(r["amount"]) > 0)
            expense = sum(-float(r["amount"]) for r in rows if float(r["amount"]) < 0)
            result.append({
                "year": y,
                "month": m,
                "label": f"{month_names_ru[m - 1]} {y}",
                "income": round(income, 2),
                "expense": round(expense, 2),
                "difference": round(income - expense, 2),
            })
    return result

# Транзакции
@app.get("/api/transactions")
async def get_transactions(
    limit: int = 100,
    month: Optional[int] = None,
    year: Optional[int] = None,
    categories: Optional[List[str]] = Query(None, alias="category"),  # мультивыбор: category=Cat1&category=Cat2
    period: Optional[List[str]] = Query(None, alias="period"),  # мультивыбор периодов: period=2025-1&period=2025-2
    type_: Optional[str] = Query(None, alias="type"),  # "income" | "expense"
    user_id: int = Depends(get_user_id)
):
    """Получить список транзакций с фильтрами (месяц, год, категория/категории, периоды, тип)"""
    db = await get_db()
    async with db.acquire() as conn:
        conditions = ["t.user_id = $1"]
        params: List = [user_id]
        n = 2
        if period and len(period) > 0:
            from datetime import date as date_type
            period_conds = []
            for p in period:
                parts = p.strip().split("-")
                if len(parts) == 2:
                    try:
                        y, m = int(parts[0]), int(parts[1])
                        if 1 <= m <= 12:
                            start = date_type(y, m, 1)
                            end = date_type(y, m + 1, 1) if m < 12 else date_type(y + 1, 1, 1)
                            period_conds.append(f"(t.created_at >= ${n}::timestamp AND t.created_at < ${n + 1}::timestamp)")
                            params.extend([start, end])
                            n += 2
                    except ValueError:
                        pass
            if period_conds:
                conditions.append("(" + " OR ".join(period_conds) + ")")
        elif month is not None and year is not None:
            from datetime import date as date_type
            start = date_type(year, month, 1)
            end = date_type(year, month + 1, 1) if month < 12 else date_type(year + 1, 1, 1)
            conditions.append(f"t.created_at >= ${n}::timestamp")
            conditions.append(f"t.created_at < ${n + 1}::timestamp")
            params.extend([start, end])
            n += 2
        cat_list = categories or []
        if cat_list:
            conditions.append(f"c.name = ANY(${n}::text[])")
            params.append(cat_list)
            n += 1
        if type_ == "income":
            conditions.append("t.amount >= 0")
        elif type_ == "expense":
            conditions.append("t.amount < 0")
        params.append(limit)
        q = f"""
            SELECT t.id, t.amount, c.name AS category, t.description, t.created_at
            FROM transactions t
            JOIN categories c ON c.id = t.category_id
            WHERE {" AND ".join(conditions)}
            ORDER BY t.created_at DESC
            LIMIT ${n}
            """
        rows = await conn.fetch(q, *params)
        return [_row_to_dict(r) for r in rows]


TRANSFER_CATEGORIES = ("Переводы людям", "Переводы от людей")


@app.get("/api/transactions/summary")
async def get_transactions_summary(
    month: Optional[int] = None,
    year: Optional[int] = None,
    categories: Optional[List[str]] = Query(None, alias="category"),
    period: Optional[List[str]] = Query(None, alias="period"),
    type_: Optional[str] = Query(None, alias="type"),
    exclude_transfers: bool = Query(False, alias="excludeTransfers"),
    user_id: int = Depends(get_user_id),
):
    """Сводка по транзакциям (суммы и количество) без лимита — для карточек Расходы/Доходы."""
    db = await get_db()
    async with db.acquire() as conn:
        conditions = ["t.user_id = $1"]
        params: List = [user_id]
        n = 2
        if period and len(period) > 0:
            from datetime import date as date_type
            period_conds = []
            for p in period:
                parts = p.strip().split("-")
                if len(parts) == 2:
                    try:
                        y, m = int(parts[0]), int(parts[1])
                        if 1 <= m <= 12:
                            start = date_type(y, m, 1)
                            end = date_type(y, m + 1, 1) if m < 12 else date_type(y + 1, 1, 1)
                            period_conds.append(f"(t.created_at >= ${n}::timestamp AND t.created_at < ${n + 1}::timestamp)")
                            params.extend([start, end])
                            n += 2
                    except ValueError:
                        pass
            if period_conds:
                conditions.append("(" + " OR ".join(period_conds) + ")")
        elif month is not None and year is not None:
            from datetime import date as date_type
            start = date_type(year, month, 1)
            end = date_type(year, month + 1, 1) if month < 12 else date_type(year + 1, 1, 1)
            conditions.append(f"t.created_at >= ${n}::timestamp")
            conditions.append(f"t.created_at < ${n + 1}::timestamp")
            params.extend([start, end])
            n += 2
        cat_list_summary = categories or []
        if cat_list_summary:
            conditions.append(f"c.name = ANY(${n}::text[])")
            params.append(cat_list_summary)
            n += 1
        if exclude_transfers:
            conditions.append(f"c.name NOT IN (${n}, ${n + 1})")
            params.extend(TRANSFER_CATEGORIES)
            n += 2
        if type_ == "income":
            conditions.append("t.amount >= 0")
        elif type_ == "expense":
            conditions.append("t.amount < 0")
        where_sql = " AND ".join(conditions)
        q = f"""
            SELECT
                COALESCE(SUM(CASE WHEN t.amount < 0 THEN -t.amount ELSE 0 END), 0) AS total_expense,
                COALESCE(SUM(CASE WHEN t.amount > 0 THEN t.amount ELSE 0 END), 0) AS total_income,
                COUNT(CASE WHEN t.amount < 0 THEN 1 END)::int AS count_expense,
                COUNT(CASE WHEN t.amount > 0 THEN 1 END)::int AS count_income
            FROM transactions t
            JOIN categories c ON c.id = t.category_id
            WHERE {where_sql}
            """
        row = await conn.fetchrow(q, *params)
        return {
            "total_expense": float(row["total_expense"]),
            "total_income": float(row["total_income"]),
            "count_expense": row["count_expense"],
            "count_income": row["count_income"],
        }


@app.post("/api/transactions")
async def create_transaction(transaction: TransactionCreate, user_id: int = Depends(get_user_id)):
    """Создать транзакцию (category_id или category по имени)"""
    db = await get_db()
    async with db.acquire() as conn:
        cid = transaction.category_id
        if cid is None and transaction.category:
            cid = await _get_category_id_by_name(conn, transaction.category)
        if cid is None:
            raise HTTPException(status_code=400, detail="Укажите category_id или category")
        await conn.execute(
            """
            INSERT INTO transactions (user_id, amount, category_id, description, created_at)
            VALUES ($1, $2, $3, $4, NOW())
            """,
            user_id, transaction.amount, cid, transaction.description
        )
        return {"status": "ok"}

@app.put("/api/transactions/{tx_id}")
async def update_transaction(
    tx_id: int,
    body: TransactionUpdate,
    user_id: int = Depends(get_user_id)
):
    """Редактировать транзакцию"""
    db = await get_db()
    async with db.acquire() as conn:
        if body.amount is not None:
            await conn.execute(
                "UPDATE transactions SET amount=$1 WHERE id=$2 AND user_id=$3",
                body.amount, tx_id, user_id
            )
        if body.category_id is not None:
            await conn.execute(
                "UPDATE transactions SET category_id=$1 WHERE id=$2 AND user_id=$3",
                body.category_id, tx_id, user_id
            )
        elif body.category is not None:
            cid = await _get_category_id_by_name(conn, body.category)
            if cid is not None:
                await conn.execute(
                    "UPDATE transactions SET category_id=$1 WHERE id=$2 AND user_id=$3",
                    cid, tx_id, user_id
                )
        if body.description is not None:
            await conn.execute(
                "UPDATE transactions SET description=$1 WHERE id=$2 AND user_id=$3",
                body.description, tx_id, user_id
            )
        return {"status": "ok"}

@app.delete("/api/transactions/{tx_id}")
async def delete_transaction(tx_id: int, user_id: int = Depends(get_user_id)):
    """Удалить транзакцию"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "DELETE FROM transactions WHERE id=$1 AND user_id=$2",
            tx_id, user_id
        )
        return {"status": "ok"}


# --- Импорт транзакций из файла (PDF, Excel, изображения) ---
# Варианты: 1) Excel/PDF — структурированный парсинг без ИИ (экономия токенов, стабильность).
#           2) PDF — regex по строкам выписки; категория через ключевые слова (_normalize_category_from_ai).
#           3) Fallback — полный парсинг через ИИ (чанки по 20k символов).

import re


def _extract_text_from_file(file_path: str, content_type: str, filename: str) -> str:
    """Извлечь текст из файла для передачи в AI-парсер."""
    ext = (filename or "").lower().split(".")[-1] if "." in (filename or "") else ""
    text_parts = []

    # Excel
    if content_type and ("spreadsheet" in content_type or "excel" in content_type) or ext in ("xlsx", "xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text_parts.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(text_parts)
        except Exception as e:
            logging.warning(f"openpyxl read failed: {e}")
            return f"Таблица (ошибка чтения): {e}"

    # PDF
    if content_type and "pdf" in content_type or ext == "pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
            return "\n".join(text_parts)
        except Exception as e:
            logging.warning(f"PDF read failed: {e}")
            return f"PDF (ошибка чтения): {e}"

    # Изображения — возвращаем заглушку; для OCR нужна отдельная библиотека
    if content_type and "image" in content_type or ext in ("png", "jpg", "jpeg"):
        return "[Изображение загружено. Распознавание изображений пока не поддерживается — загрузите PDF или Excel.]"

    return "\n".join(text_parts) if text_parts else "[Не удалось извлечь текст]"


def _fallback_category_name(amount: float) -> str:
    """Минимальный fallback для парсеров PDF/AI (резолв в id делается по имени в БД)."""
    return "Прочие расходы" if amount < 0 else "Прочие доходы"


def _is_likely_auth_code(amount: float, description: str) -> bool:
    """Проверка: не подставил ли AI код авторизации вместо суммы (4–6 целых цифр, без осмысленного описания)."""
    if amount == 0:
        return True
    try:
        v = abs(amount)
        if v != int(v):  # есть копейки — скорее сумма
            return False
        # Код авторизации обычно 4–6 цифр; при этом описание пустое или «код»/«авториз»
        desc = (description or "").strip().lower()
        if 1000 <= v <= 999999:
            if len(desc) >= 5 and "код" not in desc and "авториз" not in desc and "подтвержд" not in desc:
                return False  # похоже на реальную операцию с описанием
            if len(desc) < 3:
                return True  # нет описания — подозрительно
        return False
    except Exception:
        return False


# Месяцы по-русски для парсинга даты из выгрузки Сбера ("05 мая 2025, 09:22" или "02 фев. 2025")
# Сокращения (янв, фев, май) и полные формы в родительном падеже (мая, января, февраля)
_RU_MONTHS = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "июн": 6,
    "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
    "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
    "ноября": 11, "декабря": 12,
}


async def _parse_excel_structured(file_path: str, conn) -> tuple[list[dict], list[str]]:
    """Парсинг Excel без ИИ: форматы Сбер/Т-Банк. category_id резолвится через category_mapping (БД)."""
    transactions = []
    errors = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return [], [f"Ошибка чтения Excel: {e}"]

    if not rows:
        return [], ["Файл пуст"]

    def detect_columns(header_row):
        hdr = [str(h).strip().lower() if h is not None else "" for h in header_row]
        col_date = col_amount = col_income = col_expense = col_desc = col_type = col_category = None
        for i, h in enumerate(hdr):
            if not h or h == "none":
                continue
            if h in ("дата", "date") or h == "дата операции" or h == "дата проведения" or "дата" in h:
                if col_date is None:
                    col_date = i
            elif h in ("сумма", "amount") or h == "сумма операции" or ("сумма" in h and "в валюте" not in h and "списани" not in h and "зачислен" not in h):
                col_amount = i
            elif "сумма" in h and "в валюте" in h and col_amount is None:
                col_amount = i
            elif "сумма" in h and ("списани" in h or "расход" in h):
                col_expense = i
            elif "сумма" in h and ("зачислен" in h or "приход" in h or "доход" in h):
                col_income = i
            elif "тип операции" in h or h == "тип":
                col_type = i
            elif "категория" in h:
                col_category = i
            elif "приход" in h or "доход" in h or h == "income":
                col_income = i
            elif "расход" in h or h == "expense":
                col_expense = i
            elif h in ("описание", "назначение", "операция", "опер", "description") or "описание" in h or "назначение" in h or "название организации" in h or "название" in h:
                col_desc = i
        return col_date, col_amount, col_income, col_expense, col_desc, col_type, col_category

    # Сбер: заголовки в первой строке; Т-Банк: иногда заголовки во второй строке (первая — название отчёта)
    col_date = col_amount = col_income = col_expense = col_desc = col_type = col_category = None
    data_start_row = 1
    for try_row in range(min(4, len(rows))):
        cdate, camount, cincome, cexpense, cdesc, ctype, ccat = detect_columns(rows[try_row])
        has_any = cdate is not None or camount is not None or (cincome is not None and cexpense is not None)
        if has_any:
            col_date, col_amount, col_income, col_expense, col_desc, col_type, col_category = cdate, camount, cincome, cexpense, cdesc, ctype, ccat
            data_start_row = try_row + 1
            break

    if col_date is None and col_amount is None and col_income is None and col_expense is None:
        return [], ["Не найдены колонки даты/суммы (ожидаются заголовки: дата, сумма или приход/расход, описание)"]

    def _cell(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    def _amount_cell(row: tuple, idx: int | None) -> float | None:
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            s = str(v).replace(",", ".").replace(" ", "")
            try:
                return float(s)
            except ValueError:
                return None

    def _to_date(s: str, row_idx: int) -> str | None:
        if not s or len(s) < 6:
            return None
        s = str(s).strip()
        # YYYY-MM-DD
        if re.match(r"^\d{4}-\d{2}-\d{2}", s):
            return s[:10]
        # DD.MM.YYYY
        m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
        if m:
            d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
            return f"{y}-{mo}-{d}"
        # Сбер: "05 мая 2025, 09:22" или "02 фев. 2025" (сокращение и полная форма в род. падеже)
        m_ru = re.match(r"(\d{1,2})\s+(\w+)\s*\.?\s*(\d{4})", s, re.IGNORECASE)
        if m_ru:
            day, month_word, year = m_ru.group(1), m_ru.group(2).strip().lower(), m_ru.group(3)
            month_part = month_word[:3] if len(month_word) >= 3 else month_word
            if month_part in _RU_MONTHS:
                return f"{year}-{_RU_MONTHS[month_part]:02d}-{int(day):02d}"
            if month_word in _RU_MONTHS:
                return f"{year}-{_RU_MONTHS[month_word]:02d}-{int(day):02d}"
        return None

    for idx, row in enumerate(rows[data_start_row:], start=data_start_row + 1):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        date_str = _to_date(_cell(row, col_date), idx) if col_date is not None else None
        if not date_str and col_date is not None and col_date < len(row):
            v = row[col_date]
            if hasattr(v, "strftime"):
                date_str = v.strftime("%Y-%m-%d")
        if not date_str:
            date_str = datetime.now().strftime("%Y-%m-%d")

        amount = None
        if col_amount is not None:
            amount = _amount_cell(row, col_amount)
        if amount is None and (col_income is not None or col_expense is not None):
            inc = _amount_cell(row, col_income) or 0
            exp = _amount_cell(row, col_expense) or 0
            if inc and exp:
                amount = inc - exp
            else:
                amount = inc if inc else (-exp if exp else None)

        if amount is None:
            errors.append(f"Строка {idx}: не удалось определить сумму")
            continue
        # Тип строки по колонкам (до изменения знака): для маппинга «Прочее» и др. по bank_category_type
        is_expense_row = None
        if col_type is not None:
            type_val = _cell(row, col_type).lower()
            is_expense_row = "списание" in type_val or "расход" in type_val
        elif col_expense is not None and col_income is not None:
            inc = _amount_cell(row, col_income) or 0
            exp = _amount_cell(row, col_expense) or 0
            is_expense_row = exp > 0 and inc == 0
        else:
            is_expense_row = amount < 0 if amount else None
        # Сбер: сумма в выгрузке положительная, тип операции «Списание» — делаем расход отрицательным
        if col_type is not None and amount and amount > 0:
            type_val = _cell(row, col_type).lower()
            if "списание" in type_val or "расход" in type_val:
                amount = -amount

        description = _cell(row, col_desc) if col_desc is not None else ""
        if _is_likely_auth_code(amount, description):
            if amount != 0:
                errors.append(f"Строка {idx}: пропущена (похоже на код, не сумма): {amount}")
            continue
        bank_category = _cell(row, col_category) if col_category is not None else ""
        category_id = await _resolve_bank_category_to_id(conn, bank_category, amount, is_expense_row=is_expense_row)
        transactions.append({
            "date": date_str,
            "amount": amount,
            "category_id": category_id,
            "description": description or None,
        })

    return transactions, errors


def _parse_pdf_by_regex(raw_text: str) -> tuple[list[dict], list[str]]:
    """Извлечь транзакции из текста PDF по шаблонам строк (дата + сумма + описание). Без ИИ."""
    transactions = []
    errors = []
    lines = (raw_text or "").strip().split("\n")

    # Паттерны: дата (DD.MM.YYYY или YYYY-MM-DD) + сумма (число с точкой/запятой, возможно минус/пробелы) + остаток строки
    date_amount_desc = re.compile(
        r"^(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})[\s\t]+([+-]?\s*[\d\s.,]+?)(?:\s{2,}|\t)(.*)$",
        re.MULTILINE
    )
    # Альтернатива: сумма в конце строки перед описанием
    date_desc_amount = re.compile(
        r"^(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})[\s\t]+(.+?)[\s\t]+([+-]?\s*[\d\s.,]+)\s*$",
        re.MULTILINE
    )

    def norm_amount(s: str) -> float | None:
        s = (s or "").replace(" ", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    def to_iso_date(s: str) -> str:
        s = (s or "").strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}", s):
            return s[:10]
        m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
        if m:
            return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        return s[:10] if len(s) >= 10 else ""

    seen = set()
    for line in lines:
        line = line.strip()
        if len(line) < 12:
            continue
        for pattern in (date_amount_desc, date_desc_amount):
            m = pattern.match(line)
            if not m:
                continue
            if pattern == date_amount_desc:
                date_str = to_iso_date(m.group(1))
                amount = norm_amount(m.group(2))
                desc = (m.group(3) or "").strip()
            else:
                date_str = to_iso_date(m.group(1))
                amount = norm_amount(m.group(3))
                desc = (m.group(2) or "").strip()
            if len(date_str) < 10 or amount is None:
                continue
            if _is_likely_auth_code(amount, desc):
                continue
            key = (date_str, amount, desc[:50])
            if key in seen:
                continue
            seen.add(key)
            category = _fallback_category_name(amount)
            transactions.append({
                "date": date_str,
                "amount": amount,
                "category": category,
                "description": desc or None,
            })
            break

    if not transactions and len(lines) > 3:
        errors.append("По regex не найдено ни одной строки вида «дата сумма описание». Будет использован ИИ-парсер.")
    return transactions, errors


async def _parse_single_chunk(raw_chunk: str) -> tuple[list[dict], list[str]]:
    """Распарсить один фрагмент текста выписки (один запрос к AI)."""
    prompt = (
        "Ты парсер банковской выписки. Извлеки ВСЕ транзакции из текста.\n\n"
        "ПРАВИЛА (строго):\n"
        "1) date — дата операции в формате YYYY-MM-DD.\n"
        "2) amount — ТОЛЬКО сумма операции в рублях (положительное = приход, отрицательное = расход). "
        "Сумма обычно с копейками (1234.56) или целое. НИКОГДА не подставляй: код авторизации (4–8 цифр), "
        "последние цифры карты, коды из СМС, номера счетов — это НЕ суммы!\n"
        "3) category — ОДНА из категорий по смыслу операции: Супермаркеты, Рестораны и кафе, Транспорт, "
        "Аренда жилья, Коммунальные платежи, Здоровье и красота, Развлечения, Образование, Связь, Одежда и обувь, Дом и ремонт, Питомцы, Прочие расходы, "
        "Заработная плата, Дивиденды и купоны, Прочие доходы. Определяй по названию операции (магазин → Супермаркеты, "
        "кафе → Рестораны и кафе, АЗС/такси → Транспорт, ЖКХ → Коммунальные платежи и т.д.). НЕ относи всё в Прочее — только если смысл неясен.\n"
        "4) description — текст операции ИЗ ВЫПИСКИ как есть: название магазина/получателя/отправителя (например: ООО Магнит, Перевод Иванову). "
        "НЕ путай: category — всегда одно из слов выше; description — название из выписки.\n\n"
        "Ответь ТОЛЬКО JSON-массивом, без комментариев:\n"
        '[{"date":"YYYY-MM-DD","amount":число,"category":"одна из категорий выше","description":"название из выписки"}]\n\n'
        "Текст выписки:\n" + raw_chunk
    )
    messages = [
        {"role": "system", "content": "Ты парсер банковской выписки. Отвечай только JSON-массивом объектов с полями date, amount, category, description."},
        {"role": "user", "content": prompt}
    ]
    answer = await gigachat_request(messages)
    answer = (answer or "").strip()
    json_match = re.search(r"\[[\s\S]*\]", answer)
    if not json_match:
        return [], [f"AI не вернул массив: {answer[:150]}"]
    data = json.loads(json_match.group())
    transactions = []
    errors = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            errors.append(f"Строка {i+1}: не объект")
            continue
        try:
            date_str = str(item.get("date", ""))[:10]
            raw_amount = item.get("amount", 0)
            amount = float(raw_amount)
            raw_cat = (item.get("category") or "").strip()
            raw_desc = (item.get("description") or "").strip()
            category = (raw_cat or raw_desc) or _fallback_category_name(amount)
            description = (raw_desc or raw_cat).strip() or None
            # Отсекать коды авторизации, принятые за сумму
            if _is_likely_auth_code(amount, description or ""):
                if amount != 0:
                    errors.append(f"Строка {i+1}: пропущена (похоже на код авторизации, не сумма): {amount}")
                continue
            if not date_str or len(date_str) < 10:
                errors.append(f"Строка {i+1}: некорректная дата")
                continue
            transactions.append({
                "date": date_str,
                "amount": amount,
                "category": category,
                "description": description
            })
        except (TypeError, ValueError) as e:
            errors.append(f"Строка {i+1}: {e}")
    return transactions, errors


async def _parse_transactions_with_ai(raw_text: str) -> tuple[list[dict], list[str]]:
    """Извлечь транзакции из текста выписки. Большие выписки обрабатываются по частям (чанки)."""
    if not raw_text or len(raw_text.strip()) < 10:
        return [], ["Мало данных для распознавания"]
    text = raw_text.strip()
    # Чанки по ~20k символов: меньше запросов к AI = быстрее и меньше шанс 504
    CHUNK_SIZE = 20000
    all_transactions = []
    all_errors = []
    offset = 0
    while offset < len(text):
        chunk = text[offset:offset + CHUNK_SIZE]
        if not chunk.strip():
            offset += CHUNK_SIZE
            continue
        try:
            tx_list, err_list = await _parse_single_chunk(chunk)
            all_transactions.extend(tx_list)
            all_errors.extend(err_list)
        except Exception as e:
            all_errors.append(f"Ошибка парсинга блока: {e}")
        offset += CHUNK_SIZE
    return all_transactions, all_errors


# Сообщение, когда структура Excel не совпадает с форматами Сбер/Т-Банк
IMPORT_EXCEL_STRUCTURE_MESSAGE = "Загружаемый файл должен быть из СберОнлайн или Т-Банка без изменений"


@app.post("/api/transactions/import")
async def import_transactions_file(
    file: UploadFile = File(...),
    user_id: int = Depends(get_user_id)
):
    """Загрузить только файл Excel (выгрузка Сбера или Т-Банка без изменений). Возвращает предпросмотр (transactions + errors)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")
    fn = (file.filename or "").lower()
    if not (fn.endswith(".xlsx") or fn.endswith(".xls")):
        raise HTTPException(
            status_code=400,
            detail="Разрешена загрузка только файлов Excel (.xlsx, .xls)"
        )
    suffix = ".xlsx" if fn.endswith(".xlsx") else ".xls"
    try:
        body = await file.read()
        if len(body) > 10 * 1024 * 1024:  # 10 MB
            raise HTTPException(status_code=400, detail="File too large (max 10 MB)")
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(body)
            tmp_path = tmp.name
        try:
            db = await get_db()
            async with db.acquire() as conn:
                transactions, errors = await _parse_excel_structured(tmp_path, conn)
            if not transactions:
                return {
                    "transactions": [],
                    "errors": [IMPORT_EXCEL_STRUCTURE_MESSAGE]
                }
            ids = list({t["category_id"] for t in transactions})
            db = await get_db()
            async with db.acquire() as conn:
                rows = await conn.fetch("SELECT id, name FROM categories WHERE id = ANY($1)", ids)
            id_to_name = {r["id"]: r["name"] for r in rows}
            for t in transactions:
                t["category"] = id_to_name.get(t["category_id"], "—")
            return {"transactions": transactions, "errors": errors}
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except HTTPException:
        raise
    except Exception as e:
        logging.exception("import_transactions_file")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/transactions/import/apply")
async def import_transactions_apply(
    body: ImportApplyRequest,
    user_id: int = Depends(get_user_id)
):
    """Применить импорт: add — добавить к текущим; replace — удалить транзакции за период [min_date, max_date] из файла и вставить из файла."""
    if body.mode not in ("add", "replace"):
        raise HTTPException(status_code=400, detail="mode must be 'add' or 'replace'")
    db = await get_db()
    async with db.acquire() as conn:
        if body.mode == "replace" and body.transactions:
            dates = [t.date[:10] for t in body.transactions if t.date and len(t.date) >= 10]
            if dates:
                min_d, max_d = min(dates), max(dates)
                from datetime import date as date_type
                min_date = date_type.fromisoformat(min_d)
                max_date = date_type.fromisoformat(max_d)
                await conn.execute(
                    "DELETE FROM transactions WHERE user_id = $1 AND created_at::date >= $2 AND created_at::date <= $3",
                    user_id, min_date, max_date
                )
        for t in body.transactions:
            try:
                from datetime import datetime
                dt = datetime.strptime(t.date[:10], "%Y-%m-%d")
            except (ValueError, TypeError):
                dt = datetime.now()
            await conn.execute(
                """
                INSERT INTO transactions (user_id, amount, category_id, description, created_at)
                VALUES ($1, $2, $3, $4, $5)
                """,
                user_id, t.amount, t.category_id, (t.description or "").strip() or None, dt
            )
    return {"status": "ok", "applied": len(body.transactions)}

# Ликвидные типы для расчёта current целей: активы и пассивы
_LIQUID_ASSET_TYPES = ("Депозит", "Акции", "Облигации", "Наличные", "Банковский счёт", "Криптовалюта")
_LIQUID_LIABILITY_TYPES = ("Кредит", "Займ", "Кредитная карта", "Рассрочка")


async def _get_liquid_net(conn, user_id: int) -> float:
    """Текущий ликвидный капитал: сумма ликвидных активов минус сумма ликвидных пассивов. Один и тот же для всех целей."""
    liquid_assets = await conn.fetchval(
        """
        SELECT COALESCE(SUM(v.amount), 0)
        FROM assets a
        LEFT JOIN LATERAL (
            SELECT amount FROM asset_values WHERE asset_id = a.id ORDER BY created_at DESC LIMIT 1
        ) v ON TRUE
        WHERE a.user_id = $1 AND a.type = ANY($2::text[])
        """,
        user_id, list(_LIQUID_ASSET_TYPES)
    )
    liquid_liabilities = await conn.fetchval(
        """
        SELECT COALESCE(SUM(v.amount), 0)
        FROM liabilities l
        LEFT JOIN LATERAL (
            SELECT amount FROM liability_values WHERE liability_id = l.id ORDER BY created_at DESC LIMIT 1
        ) v ON TRUE
        WHERE l.user_id = $1 AND l.type = ANY($2::text[])
        """,
        user_id, list(_LIQUID_LIABILITY_TYPES)
    )
    return float(liquid_assets or 0) - float(liquid_liabilities or 0)


# Цели
@app.get("/api/goals")
async def get_goals(user_id: int = Depends(get_user_id)):
    """Получить список целей. current для каждой цели = ликвидный капитал (ликвидные активы − ликвидные пассивы)."""
    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, title, target, description FROM goals WHERE user_id=$1 ORDER BY id",
            user_id
        )
        current = await _get_liquid_net(conn, user_id)
        return [{"id": r["id"], "title": r["title"], "target": float(r["target"]), "current": current, "description": r["description"]} for r in rows]

@app.post("/api/goals")
async def create_goal(goal: GoalCreate, user_id: int = Depends(get_user_id)):
    """Создать цель"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO goals (user_id, target, current, title, description, created_at)
            VALUES ($1, $2, 0, $3, $4, NOW())
            """,
            user_id, goal.target, goal.title, goal.description
        )
        return {"status": "ok"}

@app.delete("/api/goals/{goal_id}")
async def delete_goal(goal_id: int, user_id: int = Depends(get_user_id)):
    """Удалить цель"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "DELETE FROM goals WHERE id=$1 AND user_id=$2",
            goal_id, user_id
        )
        return {"status": "ok"}


@app.get("/api/goals/insight")
async def get_goals_insight(user_id: int = Depends(get_user_id)):
    """Ценность 3: прогресс по целям + «через N месяцев». current = ликвидный капитал (активы − пассивы)."""
    db = await get_db()
    now = datetime.now()
    since_3m = (now.replace(day=1) - timedelta(days=90)).replace(day=1)
    async with db.acquire() as conn:
        goals_rows = await conn.fetch(
            "SELECT id, title, target FROM goals WHERE user_id=$1 ORDER BY id", user_id
        )
        current = await _get_liquid_net(conn, user_id)
        tx_rows = await conn.fetch(
            "SELECT amount FROM transactions WHERE user_id=$1 AND created_at >= $2",
            user_id, since_3m
        )
    monthly_savings = 0.0
    if tx_rows:
        total_inc = sum(float(r["amount"]) for r in tx_rows if float(r["amount"]) > 0)
        total_exp = sum(-float(r["amount"]) for r in tx_rows if float(r["amount"]) < 0)
        monthly_savings = max(0, (total_inc - total_exp) / 3) if len(tx_rows) else 0
    result = []
    for g in goals_rows:
        target = float(g["target"])
        remaining = max(0, target - current)
        months_to_goal = int(remaining / monthly_savings) if monthly_savings > 0 else None
        result.append({
            "id": g["id"],
            "title": g["title"],
            "target": target,
            "current": current,
            "remaining": remaining,
            "months_to_goal": months_to_goal,
        })
    return {"goals": result, "monthly_savings": monthly_savings}


# Бюджеты по категориям (ценность 2 — не перерасходовать)
@app.get("/api/budgets")
async def get_budgets(user_id: int = Depends(get_user_id)):
    """Список лимитов по категориям"""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            rows = await conn.fetch(
                "SELECT id, category, monthly_limit FROM budgets WHERE user_id=$1 ORDER BY category",
                user_id
            )
        except asyncpg.UndefinedTableError:
            return []
    return [{"id": r["id"], "category": r["category"], "monthly_limit": float(r["monthly_limit"])} for r in rows]


@app.get("/api/budgets/status")
async def get_budgets_status(user_id: int = Depends(get_user_id)):
    """Потрачено по категориям за текущий месяц vs лимиты (для прогресс-баров и алертов)"""
    db = await get_db()
    now = datetime.now()
    since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    async with db.acquire() as conn:
        try:
            budgets = await conn.fetch(
                "SELECT id, category, monthly_limit FROM budgets WHERE user_id=$1",
                user_id
            )
        except asyncpg.UndefinedTableError:
            return []
        result = []
        for b in budgets:
            spent_row = await conn.fetchrow(
                """
                SELECT COALESCE(SUM(ABS(amount)), 0) as s
                FROM transactions
                WHERE user_id=$1 AND category=$2 AND amount < 0 AND created_at >= $3
                """,
                user_id, b["category"], since
            )
            spent = float(spent_row["s"]) if spent_row else 0
            limit = float(b["monthly_limit"])
            result.append({
                "id": b["id"],
                "category": b["category"],
                "monthly_limit": limit,
                "spent": spent,
                "percent": min(100, int(100 * spent / limit)) if limit > 0 else 0,
            })
    return result


@app.post("/api/budgets")
async def create_budget(body: BudgetCreate, user_id: int = Depends(get_user_id)):
    """Добавить лимит по категории"""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            await conn.execute(
                """
                INSERT INTO budgets (user_id, category, monthly_limit)
                VALUES ($1, $2, $3)
                ON CONFLICT (user_id, category) DO UPDATE SET monthly_limit = $3
                """,
                user_id, body.category, body.monthly_limit
            )
        except asyncpg.UndefinedTableError:
            raise HTTPException(status_code=503, detail="Run database migrations")
    return {"status": "ok"}


@app.delete("/api/budgets/{budget_id}")
async def delete_budget(budget_id: int, user_id: int = Depends(get_user_id)):
    """Удалить лимит"""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            await conn.execute("DELETE FROM budgets WHERE id=$1 AND user_id=$2", budget_id, user_id)
        except asyncpg.UndefinedTableError:
            raise HTTPException(status_code=503, detail="Run database migrations")
    return {"status": "ok"}


# Активы
@app.get("/api/assets")
async def get_assets(user_id: int = Depends(get_user_id)):
    """Получить список активов"""
    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT a.id AS asset_id, a.title, a.type, a.currency,
                   v.amount, v.created_at AS updated_at
            FROM assets a
            LEFT JOIN LATERAL (
                SELECT amount, created_at
                FROM asset_values
                WHERE asset_id = a.id
                ORDER BY created_at DESC
                LIMIT 1
            ) v ON TRUE
            WHERE a.user_id = $1 AND (v.amount IS NULL OR v.amount > 0)
            ORDER BY a.type, v.amount ASC
            """,
            user_id
        )
        return [dict(r) for r in rows]

@app.post("/api/assets")
async def create_asset(asset: AssetCreate, user_id: int = Depends(get_user_id)):
    """Создать актив"""
    db = await get_db()
    async with db.acquire() as conn:
        asset_id = await conn.fetchval(
            """
            INSERT INTO assets (user_id, type, title, currency, created_at)
            VALUES ($1, $2, $3, 'RUB', NOW())
            RETURNING id
            """,
            user_id, asset.type, asset.title
        )
        await conn.execute(
            """
            INSERT INTO asset_values (asset_id, amount, created_at)
            VALUES ($1, $2, NOW())
            """,
            asset_id, asset.amount
        )
        return {"status": "ok", "asset_id": asset_id}

@app.put("/api/assets/{asset_id}")
async def update_asset(
    asset_id: int,
    body: AssetUpdate,
    user_id: int = Depends(get_user_id)
):
    """Редактировать актив (добавляет новую запись в asset_values при изменении суммы)"""
    db = await get_db()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id FROM assets WHERE id=$1 AND user_id=$2", asset_id, user_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Asset not found")
        if body.title is not None:
            await conn.execute(
                "UPDATE assets SET title=$1 WHERE id=$2", body.title, asset_id
            )
        if body.type is not None:
            await conn.execute(
                "UPDATE assets SET type=$1 WHERE id=$2", body.type, asset_id
            )
        if body.amount is not None:
            await conn.execute(
                """
                INSERT INTO asset_values (asset_id, amount, created_at)
                VALUES ($1, $2, NOW())
                """,
                asset_id, body.amount
            )
        return {"status": "ok"}

@app.delete("/api/assets/{asset_id}")
async def delete_asset(asset_id: int, user_id: int = Depends(get_user_id)):
    """Удалить актив"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "DELETE FROM assets WHERE id=$1 AND user_id=$2",
            asset_id, user_id
        )
        return {"status": "ok"}

# Долги
@app.get("/api/liabilities")
async def get_liabilities(user_id: int = Depends(get_user_id)):
    """Получить список долгов"""
    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT l.id AS liability_id, l.title, l.type, l.currency,
                   v.amount, v.monthly_payment, v.created_at AS updated_at
            FROM liabilities l
            LEFT JOIN LATERAL (
                SELECT amount, monthly_payment, created_at
                FROM liability_values
                WHERE liability_id = l.id
                ORDER BY created_at DESC
                LIMIT 1
            ) v ON TRUE
            WHERE l.user_id = $1 AND (v.amount IS NULL OR v.amount > 0)
            ORDER BY l.type, v.amount ASC
            """,
            user_id
        )
        return [dict(r) for r in rows]

@app.post("/api/liabilities")
async def create_liability(liability: LiabilityCreate, user_id: int = Depends(get_user_id)):
    """Создать долг"""
    db = await get_db()
    async with db.acquire() as conn:
        liability_id = await conn.fetchval(
            """
            INSERT INTO liabilities (user_id, type, title, currency, created_at)
            VALUES ($1, $2, $3, 'RUB', NOW())
            RETURNING id
            """,
            user_id, liability.type, liability.title
        )
        await conn.execute(
            """
            INSERT INTO liability_values (liability_id, amount, monthly_payment, created_at)
            VALUES ($1, $2, $3, NOW())
            """,
            liability_id, liability.amount, liability.monthly_payment
        )
        return {"status": "ok", "liability_id": liability_id}

@app.put("/api/liabilities/{liability_id}")
async def update_liability(
    liability_id: int,
    body: LiabilityUpdate,
    user_id: int = Depends(get_user_id)
):
    """Редактировать долг"""
    db = await get_db()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id FROM liabilities WHERE id=$1 AND user_id=$2",
            liability_id, user_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Liability not found")
        if body.title is not None:
            await conn.execute(
                "UPDATE liabilities SET title=$1 WHERE id=$2", body.title, liability_id
            )
        if body.type is not None:
            await conn.execute(
                "UPDATE liabilities SET type=$1 WHERE id=$2", body.type, liability_id
            )
        if body.amount is not None or body.monthly_payment is not None:
            r = await conn.fetchrow(
                "SELECT amount, monthly_payment FROM liability_values WHERE liability_id=$1 ORDER BY created_at DESC LIMIT 1",
                liability_id
            )
            amt = float(r["amount"]) if r else 0
            mp = float(r.get("monthly_payment") or 0) if r else 0
            if body.amount is not None:
                amt = body.amount
            if body.monthly_payment is not None:
                mp = body.monthly_payment
            await conn.execute(
                """
                INSERT INTO liability_values (liability_id, amount, monthly_payment, created_at)
                VALUES ($1, $2, $3, NOW())
                """,
                liability_id, amt, mp
            )
        return {"status": "ok"}

@app.delete("/api/liabilities/{liability_id}")
async def delete_liability(liability_id: int, user_id: int = Depends(get_user_id)):
    """Удалить долг"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "DELETE FROM liabilities WHERE id=$1 AND user_id=$2",
            liability_id, user_id
        )
        return {"status": "ok"}


@app.get("/api/capital/summary")
async def get_capital_summary(user_id: int = Depends(get_user_id)):
    """Текущие суммы активов, пассивов и чистый капитал (на сейчас)."""
    db = await get_db()
    async with db.acquire() as conn:
        assets_rows = await conn.fetch(
            """
            SELECT a.id, COALESCE(
                (SELECT amount FROM asset_values WHERE asset_id = a.id ORDER BY created_at DESC LIMIT 1), 0
            ) as amount
            FROM assets a WHERE a.user_id = $1
            """,
            user_id
        )
        liabs_rows = await conn.fetch(
            """
            SELECT l.id, COALESCE(
                (SELECT amount FROM liability_values WHERE liability_id = l.id ORDER BY created_at DESC LIMIT 1), 0
            ) as amount
            FROM liabilities l WHERE l.user_id = $1
            """,
            user_id
        )
    total_assets = sum(float(r["amount"]) for r in assets_rows)
    total_liabilities = sum(float(r["amount"]) for r in liabs_rows)
    return {
        "assets": round(total_assets, 2),
        "liabilities": round(total_liabilities, 2),
        "net": round(total_assets - total_liabilities, 2),
    }


@app.get("/api/capital/history")
async def get_capital_history(user_id: int = Depends(get_user_id)):
    """Активы и пассивы на последнее число каждого из последних 12 месяцев."""
    from datetime import date
    now = datetime.now()
    month_names_ru = (
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    )
    result = []
    db = await get_db()
    async with db.acquire() as conn:
        for i in range(11, -1, -1):
            m = now.month - 1 - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            last_day = date(y, m + 1, 1) - timedelta(days=1) if m < 12 else date(y + 1, 1, 1) - timedelta(days=1)
            end_dt = datetime.combine(last_day, datetime.max.time().replace(microsecond=0))
            assets_rows = await conn.fetch(
                """
                SELECT a.id, COALESCE(
                    (SELECT amount FROM asset_values
                     WHERE asset_id = a.id AND created_at <= $2
                     ORDER BY created_at DESC LIMIT 1), 0
                ) as amount
                FROM assets a WHERE a.user_id = $1
                """,
                user_id, end_dt
            )
            liabs_rows = await conn.fetch(
                """
                SELECT l.id, COALESCE(
                    (SELECT amount FROM liability_values
                     WHERE liability_id = l.id AND created_at <= $2
                     ORDER BY created_at DESC LIMIT 1), 0
                ) as amount
                FROM liabilities l WHERE l.user_id = $1
                """,
                user_id, end_dt
            )
            total_assets = sum(float(r["amount"]) for r in assets_rows)
            total_liabilities = sum(float(r["amount"]) for r in liabs_rows)
            result.append({
                "year": y,
                "month": m,
                "label": f"{month_names_ru[m - 1]} {y}",
                "assets": round(total_assets, 2),
                "liabilities": round(total_liabilities, 2),
                "net": round(total_assets - total_liabilities, 2),
            })
    return result


# --- Чек-листы из консультации и цель месяца ---

@app.get("/api/consultation/actions")
async def get_consultation_actions(user_id: int = Depends(get_user_id)):
    """Список действий из консультации (чек-лист) с флагом done."""
    db = await get_db()
    try:
        async with db.acquire() as conn:
            rows = await conn.fetch(
                "SELECT id, action_text, done, created_at FROM user_consultation_actions WHERE user_id=$1 ORDER BY created_at DESC LIMIT 20",
                user_id
            )
        return [{"id": r["id"], "action_text": r["action_text"], "done": r["done"], "created_at": r["created_at"].isoformat() if r.get("created_at") else None} for r in rows]
    except asyncpg.UndefinedTableError:
        return []


@app.patch("/api/consultation/actions/{action_id}")
async def patch_consultation_action(
    action_id: int,
    body: dict,
    user_id: int = Depends(get_user_id)
):
    """Отметить действие выполненным (body: {"done": true/false})."""
    done = body.get("done")
    if done is None:
        raise HTTPException(status_code=400, detail="done required")
    try:
        db = await get_db()
        async with db.acquire() as conn:
            await conn.execute(
                "UPDATE user_consultation_actions SET done = $1 WHERE id = $2 AND user_id = $3",
                bool(done), action_id, user_id
            )
    except asyncpg.UndefinedTableError:
        raise HTTPException(status_code=404, detail="Table not found. Run migration.")
    return {"status": "ok"}


@app.get("/api/focus-goal")
async def get_focus_goal(user_id: int = Depends(get_user_id)):
    """Цель на этот месяц (фокус от ИИ)."""
    now = datetime.now()
    db = await get_db()
    try:
        async with db.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT id, title, target_amount, for_month, for_year, achieved_at FROM user_focus_goal WHERE user_id=$1 AND for_month=$2 AND for_year=$3",
                user_id, now.month, now.year
            )
        if not row:
            return None
        return {
            "id": row["id"],
            "title": row["title"],
            "target_amount": float(row["target_amount"]),
            "for_month": row["for_month"],
            "for_year": row["for_year"],
            "achieved_at": row["achieved_at"].isoformat() if row.get("achieved_at") else None,
        }
    except asyncpg.UndefinedTableError:
        return None


@app.patch("/api/focus-goal/{goal_id}")
async def patch_focus_goal_achieved(goal_id: int, user_id: int = Depends(get_user_id)):
    """Отметить цель месяца достигнутой."""
    try:
        db = await get_db()
        async with db.acquire() as conn:
            await conn.execute(
                "UPDATE user_focus_goal SET achieved_at = NOW() WHERE id = $1 AND user_id = $2",
                goal_id, user_id
            )
    except asyncpg.UndefinedTableError:
        raise HTTPException(status_code=404, detail="Table not found. Run migration.")
    return {"status": "ok"}


# --- Бенчмарки (анонимные: целевые диапазоны + доля пользователя) ---

# Целевые диапазоны % от дохода (нормы)
_BENCHMARK_TARGETS = {
    "Еда и продукты": (15, 35),
    "Развлечения": (5, 15),
    "Транспорт": (5, 15),
    "Жильё и коммуналка": (20, 35),
    "Здоровье": (5, 15),
    "Связь и интернет": (2, 5),
    "Сбережения": (10, 20),
}


# Категории дохода, облагаемого налогом (для расчёта долей расходов)
_TAXABLE_INCOME_CATEGORIES = ("Зарплата", "Инвестиции и дивиденды", "Прочие доходы")


@app.get("/api/benchmarks")
async def get_benchmarks(user_id: int = Depends(get_user_id)):
    """Бенчмарки: доля от налогооблагаемого дохода (зарплата, дивиденды, прочие доходы). Период: последний год или доступный. Показать только Сбережения и категории, превышающие целевую норму."""
    db = await get_db()
    now = datetime.now()
    since = (now.replace(day=1) - timedelta(days=365)).replace(day=1)
    async with db.acquire() as conn:
        # Доход за период: только налогооблагаемый (по категориям)
        income_row = await conn.fetchval(
            """
            SELECT COALESCE(SUM(t.amount), 0)
            FROM transactions t
            JOIN categories c ON c.id = t.category_id AND c.type = 'Доход'
            WHERE t.user_id = $1 AND t.amount > 0 AND t.created_at >= $2
            AND c.name = ANY($3::text[])
            """,
            user_id, since, list(_TAXABLE_INCOME_CATEGORIES),
        )
        total_income = float(income_row or 0)
        if total_income <= 0:
            # Fallback: весь доход за период
            income_row = await conn.fetchval(
                "SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE user_id=$1 AND amount > 0 AND created_at >= $2",
                user_id, since,
            )
            total_income = float(income_row or 0)
        if total_income <= 0:
            return {"total_income": 0, "categories": [], "savings": None, "period_months": 12}
        rows = await conn.fetch(
            """
            SELECT c.name, SUM(ABS(t.amount)) as total
            FROM transactions t
            JOIN categories c ON c.id = t.category_id
            WHERE t.user_id=$1 AND t.amount < 0 AND t.created_at >= $2
            GROUP BY c.name
            """,
            user_id, since
        )
    expenses_by_cat = {r["name"]: float(r["total"]) for r in rows}
    total_expense = sum(expenses_by_cat.values())
    savings = total_income - total_expense
    savings_pct = round((savings / total_income) * 100, 1) if total_income else 0
    out_categories = []
    for cat_name, target_range in _BENCHMARK_TARGETS.items():
        if cat_name == "Сбережения":
            continue
        exp = expenses_by_cat.get(cat_name, 0)
        pct = round((exp / total_income) * 100, 1) if total_income else 0
        # Показываем только если превышает целевую норму (user_pct > target_high)
        if pct > target_range[1]:
            out_categories.append({
                "name": cat_name,
                "user_pct": pct,
                "target_low": target_range[0],
                "target_high": target_range[1],
            })
    savings_item = {
        "name": "Сбережения",
        "user_pct": savings_pct,
        "target_low": 10,
        "target_high": 20,
    }
    return {
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "categories": out_categories,
        "savings": savings_item,
        "period_months": 12,
    }


# --- Прогресс относительно себя ---

@app.get("/api/progress-vs-self")
async def get_progress_vs_self(user_id: int = Depends(get_user_id)):
    """Сравнение: среднее в месяц за последние 12 месяцев vs последний месяц по категориям расходов. Топ-3 по разнице."""
    db = await get_db()
    now = datetime.now()
    # Последние 12 месяцев (для среднего в месяц)
    start_12 = (now.replace(day=1) - timedelta(days=365)).replace(day=1)
    end_12 = now
    # Последний календарный месяц
    first_day_this = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day_prev = first_day_this - timedelta(days=1)
    start_last_month = last_day_prev.replace(day=1)
    end_last_month = last_day_prev
    async with db.acquire() as conn:
        # Суммы по категориям за 12 месяцев
        rows_12 = await conn.fetch(
            """
            SELECT c.name, SUM(ABS(t.amount)) as total
            FROM transactions t JOIN categories c ON c.id = t.category_id
            WHERE t.user_id=$1 AND t.amount < 0 AND t.created_at >= $2 AND t.created_at <= $3
            GROUP BY c.name
            """,
            user_id, start_12, end_12
        )
        # Суммы по категориям за последний месяц
        rows_last = await conn.fetch(
            """
            SELECT c.name, SUM(ABS(t.amount)) as total
            FROM transactions t JOIN categories c ON c.id = t.category_id
            WHERE t.user_id=$1 AND t.amount < 0 AND t.created_at >= $2 AND t.created_at <= $3
            GROUP BY c.name
            """,
            user_id, start_last_month, end_last_month
        )
    # Исключаем категории «переводы» (как в фильтре транзакций)
    _PROGRESS_EXCLUDE_CATEGORIES = ("Переводы людям", "Переводы от людей")
    by_cat_12 = {r["name"]: float(r["total"]) for r in rows_12 if r["name"] not in _PROGRESS_EXCLUDE_CATEGORIES}
    by_cat_last = {r["name"]: float(r["total"]) for r in rows_last if r["name"] not in _PROGRESS_EXCLUDE_CATEGORIES}
    all_cats = set(by_cat_12) | set(by_cat_last)
    months_with_data = 12
    result = []
    for cat in all_cats:
        total_12 = by_cat_12.get(cat, 0)
        last_val = by_cat_last.get(cat, 0)
        avg_per_month = total_12 / months_with_data if months_with_data else 0
        diff = last_val - avg_per_month
        result.append({
            "category": cat,
            "before": round(avg_per_month, 2),
            "now": round(last_val, 2),
            "diff": round(diff, 2),
        })
    # Топ-3 по величине разницы (не по модулю: сначала самые большие положительные, затем отрицательные)
    result.sort(key=lambda x: x["diff"], reverse=True)
    result = result[:3]
    return {
        "period_before": "в ср. в мес. (12 мес.)",
        "period_now": "последний месяц",
        "categories": result,
    }


# --- Прогресс онбординга (пайплайн) ---

@app.get("/api/onboarding-progress")
async def get_onboarding_progress(user_id: int = Depends(get_user_id)):
    """Флаги для пайплайна онбординга: 1 транзакция, 1 актив/долг, профиль заполнен, 1 консультация."""
    db = await get_db()
    async with db.acquire() as conn:
        has_tx = await conn.fetchval(
            "SELECT 1 FROM transactions WHERE user_id = $1 LIMIT 1", user_id
        )
        has_assets = await conn.fetchval("SELECT 1 FROM assets WHERE user_id = $1 LIMIT 1", user_id)
        has_liabs = await conn.fetchval("SELECT 1 FROM liabilities WHERE user_id = $1 LIMIT 1", user_id)
        profile_row = await conn.fetchrow(
            "SELECT gender, birth_date, marital_status, children_count, city FROM users WHERE id = $1",
            user_id,
        )
        has_profile = False
        if profile_row:
            has_profile = bool(
                profile_row.get("gender")
                or profile_row.get("birth_date")
                or profile_row.get("marital_status")
                or profile_row.get("children_count") is not None
                or profile_row.get("city")
            )
        has_consultation = await conn.fetchval(
            """
            SELECT 1 FROM ai_context
            WHERE user_id = $1 AND role = 'assistant' AND content LIKE 'CONSULTATION:%' LIMIT 1
            """,
            user_id,
        )
    return {
        "has_transactions": bool(has_tx),
        "has_capital": bool(has_assets or has_liabs),
        "has_profile": has_profile,
        "has_consultation": bool(has_consultation),
    }


# --- Мягкие алерты ---

@app.get("/api/alerts")
async def get_alerts(user_id: int = Depends(get_user_id)):
    """Мягкие алерты: расходы выше обычного, до цели по резерву и т.д."""
    db = await get_db()
    now = datetime.now()
    alerts = []
    async with db.acquire() as conn:
        # Расходы текущего месяца vs средние за предыдущие 3
        cur_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        cur_exp = await conn.fetchval(
            "SELECT COALESCE(SUM(ABS(amount)), 0) FROM transactions WHERE user_id=$1 AND amount < 0 AND created_at >= $2",
            user_id, cur_month_start
        )
        prev_start = (cur_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=90)
        prev_end = cur_month_start - timedelta(days=1)
        prev_avg = await conn.fetchval(
            "SELECT COALESCE(SUM(ABS(amount)), 0) / 3 FROM transactions WHERE user_id=$1 AND amount < 0 AND created_at >= $2 AND created_at <= $3",
            user_id, prev_start, prev_end
        )
        cur_exp = float(cur_exp or 0)
        prev_avg = float(prev_avg or 0)
        if prev_avg > 0 and cur_exp > prev_avg * 1.15:
            pct = int(round((cur_exp / prev_avg - 1) * 100))
            alerts.append({"type": "expense_above", "text": f"В этом месяце расходы пока на {pct}% выше среднего за предыдущие 3 месяца."})
        # Резервный фонд: месячные расходы и сколько месяцев покрывает ликвидный капитал
        liquid_net = await _get_liquid_net(conn, user_id)
        if prev_avg > 0 and liquid_net >= 0:
            months_reserve = liquid_net / prev_avg
            if months_reserve < 3:
                alerts.append({"type": "reserve_low", "text": f"Резервный фонд покрывает около {months_reserve:.1f} мес. расходов. Рекомендуется 3–6 мес."})
            elif months_reserve >= 3:
                alerts.append({"type": "reserve_ok", "text": f"Резервный фонд покрывает около {months_reserve:.1f} мес. расходов."})
    return {"alerts": alerts}


# --- Симулятор сценариев ---

@app.get("/api/simulator")
async def get_simulator(
    goal_id: Optional[int] = Query(None),
    monthly_savings: Optional[float] = Query(None),
    monthly_payment: Optional[float] = Query(None),
    user_id: int = Depends(get_user_id)
):
    """Сценарии: до цели при откладывании X в месяц; до погашения долга при платеже Y в месяц."""
    db = await get_db()
    result = {"goal_months": None, "debt_months": None}
    async with db.acquire() as conn:
        if goal_id is not None and monthly_savings is not None and monthly_savings > 0:
            row = await conn.fetchrow("SELECT target, current FROM goals WHERE id=$1 AND user_id=$2", goal_id, user_id)
            if row:
                remaining = float(row["target"]) - float(row["current"])
                if remaining > 0:
                    months = max(1, int(remaining / monthly_savings))
                    result["goal_months"] = months
        # Долг: сумма пассивов; при monthly_payment — сколько месяцев до нуля
        if monthly_payment is not None and monthly_payment > 0:
            total_debt = await conn.fetchval(
                """
                SELECT COALESCE(SUM(v.amount), 0) FROM liabilities l
                LEFT JOIN LATERAL (SELECT amount FROM liability_values WHERE liability_id = l.id ORDER BY created_at DESC LIMIT 1) v ON TRUE
                WHERE l.user_id = $1
                """,
                user_id
            )
            total_debt = float(total_debt or 0)
            if total_debt > 0:
                result["debt_months"] = max(1, int(total_debt / monthly_payment))
                result["total_debt"] = total_debt
    return result


# --- Отметки прогресса (бейджи) ---

@app.get("/api/badges")
async def get_badges(user_id: int = Depends(get_user_id)):
    """Бейджи: резервный фонд на N мес., первая цель достигнута и т.д."""
    db = await get_db()
    badges = []
    async with db.acquire() as conn:
        liquid_net = await _get_liquid_net(conn, user_id)
        avg_expense = await conn.fetchval(
            "SELECT COALESCE(SUM(ABS(amount)), 0) / 3 FROM transactions WHERE user_id=$1 AND amount < 0 AND created_at >= $2",
            user_id, (datetime.now() - timedelta(days=120)).replace(day=1)
        )
        avg_expense = float(avg_expense or 0)
        if avg_expense > 0 and liquid_net >= 0:
            months_reserve = liquid_net / avg_expense
            if months_reserve >= 3:
                badges.append({"id": "reserve_3", "label": f"Резервный фонд на {int(months_reserve)} мес."})
            if months_reserve >= 6:
                badges.append({"id": "reserve_6", "label": "Резервный фонд на 6+ мес."})
        goals_done = await conn.fetchval(
            "SELECT COUNT(*) FROM goals WHERE user_id=$1 AND current >= target AND target > 0",
            user_id
        )
        if goals_done and int(goals_done) > 0:
            badges.append({"id": "first_goal", "label": "Первая цель достигнута"})
    return {"badges": badges}


# ============================================
# AI Functions (GigaChat)
# ============================================

async def get_gigachat_token():
    """Получить токен доступа GigaChat"""
    auth_str = f"{G_CLIENT_ID}:{G_CLIENT_SECRET}"
    b64 = base64.b64encode(auth_str.encode()).decode()
    headers = {
        "Authorization": f"Basic {b64}",
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json",
        "RqUID": str(uuid.uuid4())
    }
    data = {"scope": G_SCOPE}
    async with httpx.AsyncClient(verify=False, timeout=20.0) as client:
        r = await client.post(G_AUTH_URL, headers=headers, data=data)
        r.raise_for_status()
        return r.json().get("access_token")


async def gigachat_request(messages):
    """Запрос к GigaChat API"""
    token = await get_gigachat_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    payload = {
        "model": GIGACHAT_MODEL,
        "messages": messages,
        "temperature": 0.3
    }
    async with httpx.AsyncClient(verify=False, timeout=40.0) as client:
        r = await client.post(G_API_URL, headers=headers, json=payload)
        r.raise_for_status()
        j = r.json()
        if "choices" in j and j["choices"]:
            return j["choices"][0]["message"]["content"]
        return json.dumps(j, ensure_ascii=False)


# AI Cache helpers
def _hash_input(user_message: str, finance_snapshot: str) -> str:
    """Хеширование входных данных для кэша"""
    h = hashlib.sha256((user_message.strip().lower() + "\n" + finance_snapshot).encode("utf-8"))
    return h.hexdigest()


async def get_cached_ai_reply(user_id: int, user_message: str, finance_snapshot: str):
    """Получить ответ из кэша"""
    h = _hash_input(user_message, finance_snapshot)
    db = await get_db()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT answer FROM ai_cache WHERE user_id=$1 AND input_hash=$2 ORDER BY created_at DESC LIMIT 1",
            user_id, h
        )
        return row["answer"] if row else None


async def save_ai_cache(user_id: int, user_message: str, finance_snapshot: str, ai_answer: str):
    """Сохранить ответ в кэш"""
    h = _hash_input(user_message, finance_snapshot)
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "INSERT INTO ai_cache (user_id, input_hash, answer, created_at) VALUES ($1,$2,$3,NOW())",
            user_id, h, ai_answer
        )


async def save_message(user_id: int, role: str, content: str):
    """Сохранить сообщение в контекст"""
    db = await get_db()
    async with db.acquire() as conn:
        await conn.execute(
            "INSERT INTO ai_context (user_id, role, content, created_at) VALUES ($1,$2,$3,NOW())",
            user_id, role, content
        )


async def analyze_user_finances_text(user_id: int) -> str:
    """Анализ финансов пользователя для AI (включая профиль, если есть)."""
    MAX_TX_FOR_ANALYSIS = 200
    db = await get_db()
    async with db.acquire() as conn:
        # Профиль пользователя (пол, возраст, семья, город) — для персональных рекомендаций
        profile_lines = []
        try:
            profile_row = await conn.fetchrow(
                "SELECT gender, birth_date, marital_status, children_count, city FROM users WHERE id=$1",
                user_id
            )
            if profile_row and (profile_row.get("gender") or profile_row.get("birth_date") or profile_row.get("marital_status") or profile_row.get("children_count") is not None or profile_row.get("city")):
                if profile_row.get("gender"):
                    profile_lines.append(f"Пол: {profile_row['gender']}")
                if profile_row.get("birth_date"):
                    from datetime import date
                    bd = profile_row["birth_date"]
                    age = (date.today() - bd).days // 365 if isinstance(bd, date) else None
                    if age is not None:
                        profile_lines.append(f"Возраст: {age} лет (дата рождения: {bd})")
                if profile_row.get("marital_status"):
                    profile_lines.append(f"Семейное положение: {profile_row['marital_status']}")
                if profile_row.get("children_count") is not None:
                    profile_lines.append(f"Дети: {profile_row['children_count']}")
                if profile_row.get("city"):
                    profile_lines.append(f"Город: {profile_row['city']}")
        except asyncpg.UndefinedColumnError:
            pass
        if profile_lines:
            s = "Профиль пользователя:\n" + "\n".join(profile_lines) + "\n\n"
        else:
            s = ""

        rows = await conn.fetch(
            """
            SELECT t.amount, c.name AS category, t.description, t.created_at
            FROM transactions t
            LEFT JOIN categories c ON c.id = t.category_id
            WHERE t.user_id=$1 ORDER BY t.created_at DESC LIMIT $2
            """,
            user_id, MAX_TX_FOR_ANALYSIS
        )
        
        if rows:
            s += "Последние транзакции:\n"
            for r in rows:
                ts = r["created_at"].strftime("%Y-%m-%d") if r["created_at"] else ""
                s += f"- {r['amount']}₽ | {r.get('category') or '-'} | {r.get('description') or ''} | {ts}\n"
        else:
            s += "У пользователя нет транзакций.\n"
        
        goals = await conn.fetch("SELECT title, target FROM goals WHERE user_id=$1", user_id)
        if goals:
            liquid_net = await _get_liquid_net(conn, user_id)
            s += "\nЦели:\n"
            for g in goals:
                s += f"- {g.get('title','Цель')}: {liquid_net}/{g['target']} ₽\n"
        
        # Активы
        assets_rows = await conn.fetch(
            """
            SELECT a.title, a.type, v.amount
            FROM assets a
            LEFT JOIN LATERAL (
                SELECT amount FROM asset_values WHERE asset_id = a.id ORDER BY created_at DESC LIMIT 1
            ) v ON TRUE
            WHERE a.user_id = $1 AND (v.amount IS NULL OR v.amount > 0)
            """,
            user_id
        )
        if assets_rows:
            total_assets = sum([a["amount"] for a in assets_rows if a["amount"]])
            s += f"\nАктивы (итого {total_assets}₽):\n"
            for a in assets_rows:
                if a["amount"]:
                    s += f"- {a['title']} ({a['type']}): {a['amount']}₽\n"
        
        # Долги
        liabs_rows = await conn.fetch(
            """
            SELECT l.title, l.type, v.amount
            FROM liabilities l
            LEFT JOIN LATERAL (
                SELECT amount FROM liability_values WHERE liability_id = l.id ORDER BY created_at DESC LIMIT 1
            ) v ON TRUE
            WHERE l.user_id = $1 AND (v.amount IS NULL OR v.amount > 0)
            """,
            user_id
        )
        if liabs_rows:
            total_liabs = sum([l["amount"] for l in liabs_rows if l["amount"]])
            s += f"\nДолги (итого {total_liabs}₽):\n"
            for l in liabs_rows:
                if l["amount"]:
                    s += f"- {l['title']} ({l['type']}): {l['amount']}₽\n"
        
        total_assets = sum([a["amount"] for a in assets_rows if a.get("amount")]) if assets_rows else 0
        total_liabs = sum([l["amount"] for l in liabs_rows if l.get("amount")]) if liabs_rows else 0
        s += f"\nЧистый капитал: {total_assets - total_liabs}₽\n"
        # Выполненные действия из прошлых консультаций (для учёта в новой)
        try:
            done_actions = await conn.fetch(
                "SELECT action_text FROM user_consultation_actions WHERE user_id=$1 AND done=TRUE ORDER BY created_at DESC LIMIT 10",
                user_id
            )
            if done_actions:
                s += "\nВыполненные действия пользователя (уже сделано):\n"
                for a in done_actions:
                    s += f"- {a['action_text']}\n"
        except asyncpg.UndefinedTableError:
            pass
        return s


def _stub_consultation_text(finance_snapshot: str) -> str:
    """Заглушка консультации для тестовой среды без GigaChat."""
    return (
        "📊 *Ваша финансовая консультация (тестовый режим)*\n\n"
        "Данные для анализа:\n"
        f"{finance_snapshot[:800]}{'…' if len(finance_snapshot) > 800 else ''}\n\n"
        "💰 *Рекомендации*\n"
        "• В продакшене здесь будет персональный отчёт от ИИ.\n"
        "• Настройте GIGACHAT_* в .env для полной генерации."
    )


async def generate_consultation(user_id: int) -> str:
    """Генерация финансовой консультации"""
    try:
        finance_snapshot = await analyze_user_finances_text(user_id)

        # Тестовая среда без GigaChat: возвращаем заглушку, чтобы генерация «работала»
        if APP_ENV == "test" and not (G_CLIENT_ID and G_API_URL):
            stub = _stub_consultation_text(finance_snapshot or "Нет данных.")
            await save_message(user_id, "assistant", f"CONSULTATION: {stub}")
            return stub

        # Если нет данных
        if not finance_snapshot or ("нет транзакций" in finance_snapshot.lower() and "нет активов" in finance_snapshot.lower()):
            return (
                "📊 *Ваша финансовая консультация*\n\n"
                "У вас пока нет финансовых данных для анализа.\n\n"
                "Рекомендации для начала:\n"
                "1. Начните вести учет доходов и расходов\n"
                "2. Добавьте информацию о ваших активах\n"
                "3. Установите финансовые цели\n"
                "4. Регулярно обновляйте данные\n\n"
                "После добавления данных вы получите персональные рекомендации!"
            )
        
        # Проверяем кэш
        cached = await get_cached_ai_reply(user_id, "CONSULT_REQUEST", finance_snapshot)
        if cached:
            return cached
        
        system_prompt = (
            "Ты — опытный персональный финансовый консультант. Твой тон: профессионально, понятно, без морализаторства. "
            "Пиши так, как будто консультируешь живого человека. Не используй сложные термины без пояснений.\n\n"
            "ОБЯЗАТЕЛЬНО используй в анализе:\n"
            "1. ТРАНЗАКЦИИ — доходы и расходы, паттерны, топ категорий с суммами.\n"
            "2. ЦЕЛИ — цели пользователя и прогресс по ним (если есть).\n"
            "3. АКТИВЫ и ДОЛГИ — структура капитала, влияние на бюджет.\n"
            "4. ПРОФИЛЬ (если есть) — возраст, семья, город — учитывай в рекомендациях.\n\n"
            "ФОРМАТ ОТВЕТА (соблюдай структуру):\n\n"
            "1. **Краткий вывод (1–3 абзаца)**\n"
            "Дай общее резюме финансового состояния:\n"
            "- устойчивость\n"
            "- баланс доходов и расходов\n"
            "- уровень риска\n"
            "- сильные и слабые стороны\n\n"
            "2. **Анализ текущей ситуации**\n"
            "Разбери по пунктам:\n"
            "- денежный поток (доходы vs расходы)\n"
            "- наличие или отсутствие свободного остатка\n"
            "- долговую нагрузку (оценка адекватности)\n"
            "- подушку безопасности (в месяцах расходов)\n"
            "- структуру капитала (ликвидность, концентрация рисков)\n\n"
            "3. **Проблемные зоны и риски**\n"
            "Укажи:\n"
            "- где пользователь теряет деньги или время\n"
            "- какие риски могут привести к финансовым проблемам\n"
            "- что требует внимания в первую очередь\n\n"
            "4. **Конкретные рекомендации (по приоритету)**\n"
            "Дай пошаговый план:\n"
            "- что сделать в ближайшие 1–3 месяца\n"
            "- что делать в горизонте 6–12 месяцев\n"
            "- что улучшать в долгосрочной перспективе\n\n"
            "Рекомендации должны быть:\n"
            "- конкретными (с цифрами, если возможно)\n"
            "- реалистичными\n"
            "- адаптированными под город и жизненную ситуацию пользователя\n\n"
            "5. **Работа с целями**\n"
            "- оцени реалистичность целей\n"
            "- предложи стратегию их достижения\n"
            "- если целей нет — предложи сформулировать 2–3 базовые цели\n\n"
            "6. **Тон и стиль**\n"
            "- Профессионально, понятно, без морализаторства\n"
            "- Не использовать сложные термины без пояснений\n"
            "- Писать так, как будто консультируешь живого человека\n\n"
            "ОГРАНИЧЕНИЯ:\n"
            "- Не давай юридических или налоговых советов\n"
            "- Не гарантируй доходность\n"
            "- Не используй фразы вроде «вам обязательно нужно», вместо этого — «рационально рассмотреть», «оптимальным шагом будет»\n\n"
            "В конце консультации задай 1–2 уточняющих вопроса, которые помогут дать ещё более точные рекомендации в будущем.\n\n"
            "ТРЕБОВАНИЯ К ФОРМАТУ:\n"
            "- Используй **текст** для заголовков разделов (двойные звездочки)\n"
            "- Используй *текст* для подзаголовков и выделения (одинарные звездочки)\n"
            "- Конкретные суммы везде с пробелами: 200 000 ₽, 1 500 000 ₽\n"
            "- Не общие фразы («пересмотреть») — только конкретика с цифрами\n"
            "- Не дублируй информацию из других отчетов (пользователь уже видит цифры в дашборде)\n"
            "- Проанализируй ответы по промпту и улучши их, чтобы пользователь понимал ценность рекомендаций\n\n"
            "🚨 ФОРМАТ ЧИСЕЛ (строго):\n"
            "- Только с пробелами: 200 000 ₽, 1 500 000 ₽. Запрещены: научная нотация (2.7E+5), "
            "точки как разделители (12.000.000), дробная часть для рублей. Округляй до целых.\n\n"
            "Отвечай на русском.\n\n"
            "В конце ответа добавь блок (если ещё не включил в план):\n"
            "CHECK: [действие 1 с суммой]\n"
            "CHECK: [действие 2]\n"
            "CHECK: [действие 3]\n"
            "FOCUS_MONTH: [одна цель на этот месяц, например: Отложить 5 000 ₽]\n"
            "(CHECK — 2–3 конкретных действия с галочкой для пользователя; FOCUS_MONTH — одна фокус-цель на текущий месяц.)"
        )
        messages = [
            {"role":"system","content":system_prompt},
            {"role":"user","content":finance_snapshot}
        ]
        
        answer = await gigachat_request(messages)
        
        if not answer or len(answer.strip()) == 0:
            return "Извините, не удалось сгенерировать консультацию. Попробуйте позже."
        
        await save_message(user_id, "assistant", f"CONSULTATION: {answer}")
        await save_ai_cache(user_id, "CONSULT_REQUEST", finance_snapshot, answer)

        # Парсим CHECK: и FOCUS_MONTH из ответа, сохраняем в БД
        import re
        now = datetime.now()
        for line in answer.split("\n"):
            line = line.strip()
            if line.upper().startswith("CHECK:"):
                action_text = line[6:].strip()
                if action_text:
                    try:
                        db = await get_db()
                        async with db.acquire() as conn:
                            await conn.execute(
                                "INSERT INTO user_consultation_actions (user_id, action_text, done) VALUES ($1, $2, FALSE)",
                                user_id, action_text
                            )
                    except asyncpg.UndefinedTableError:
                        pass
            if line.upper().startswith("FOCUS_MONTH:"):
                focus_text = line[11:].strip()
                if focus_text:
                    # Пытаемся извлечь сумму из текста (числа с пробелами или без)
                    nums = re.findall(r"[\d\s]+", focus_text)
                    target_amount = 0
                    for n in nums:
                        clean = int("".join(n.split()))
                        if clean > target_amount:
                            target_amount = clean
                    if target_amount == 0:
                        target_amount = 1
                    try:
                        db = await get_db()
                        async with db.acquire() as conn:
                            await conn.execute(
                                """
                                INSERT INTO user_focus_goal (user_id, title, target_amount, for_month, for_year)
                                VALUES ($1, $2, $3, $4, $5)
                                ON CONFLICT (user_id, for_month, for_year) DO UPDATE SET title = $2, target_amount = $3
                                """,
                                user_id, focus_text[:500], target_amount, now.month, now.year
                            )
                    except asyncpg.UndefinedTableError:
                        pass

        return answer
        
    except Exception as e:
        logging.error(f"Ошибка при генерации консультации: {e}")
        import traceback
        traceback.print_exc()
        return (
            "❌ *Ошибка при генерации консультации*\n\n"
            "Извините, произошла техническая ошибка.\n"
            "Попробуйте позже или обратитесь в поддержку."
        )


# Отчеты
@app.get("/api/reports")
async def get_reports(user_id: int = Depends(get_user_id)):
    """Получить отчеты (3 графика с пояснениями)"""
    db = await get_db()
    async with db.acquire() as conn:
        # График 1: Расходы по категориям за текущий месяц
        now = datetime.now()
        since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        expense_rows = await conn.fetch(
            """
            SELECT category, SUM(ABS(amount)) as total
            FROM transactions
            WHERE user_id=$1 AND created_at >= $2 AND amount < 0
            GROUP BY category
            ORDER BY total DESC
            LIMIT 10
            """,
            user_id, since
        )
        
        expense_by_cat = {r['category'] or 'Прочее': float(r['total']) for r in expense_rows}
        total_expenses = sum(expense_by_cat.values())
        
        # График 2: Прогресс по целям
        goals_rows = await conn.fetch(
            """
            SELECT title, target FROM goals WHERE user_id=$1 ORDER BY id
            """,
            user_id
        )
        liquid_net = await _get_liquid_net(conn, user_id)
        def _goal_progress(current: float, target: float) -> float:
            """Прогресс 0–100%; при target=0 возвращаем 100%; деление на 0 исключено."""
            if target <= 0:
                return 100.0
            p = (max(0, current) / target) * 100
            return max(0.0, min(100.0, p))
        goals_data = [
            {
                'title': g['title'],
                'target': float(g['target']),
                'current': liquid_net,
                'progress': _goal_progress(liquid_net, float(g['target']))
            }
            for g in goals_rows
        ]
        
        # График 3: Динамика капитала за последние 12 недель
        weeks_data = []
        for i in range(11, -1, -1):
            week_end = now - timedelta(weeks=i)
            # Находим воскресенье недели
            days_since_monday = (week_end.weekday()) % 7
            sunday = week_end - timedelta(days=days_since_monday) + timedelta(days=6)
            sunday = sunday.replace(hour=23, minute=59, second=59)
            
            # Получаем активы на эту дату
            assets_rows = await conn.fetch(
                """
                SELECT a.id, COALESCE(
                    (SELECT amount FROM asset_values 
                     WHERE asset_id = a.id AND created_at <= $2 
                     ORDER BY created_at DESC LIMIT 1), 0
                ) as amount
                FROM assets a
                WHERE a.user_id = $1
                """,
                user_id, sunday
            )
            total_assets = sum(float(r['amount']) for r in assets_rows)
            
            # Получаем долги на эту дату
            liabs_rows = await conn.fetch(
                """
                SELECT l.id, COALESCE(
                    (SELECT amount FROM liability_values 
                     WHERE liability_id = l.id AND created_at <= $2 
                     ORDER BY created_at DESC LIMIT 1), 0
                ) as amount
                FROM liabilities l
                WHERE l.user_id = $1
                """,
                user_id, sunday
            )
            total_liabs = sum(float(r['amount']) for r in liabs_rows)
            
            net_capital = total_assets - total_liabs
            weeks_data.append({
                'week': sunday.strftime('%d.%m'),
                'assets': total_assets,
                'liabilities': total_liabs,
                'net_capital': net_capital
            })
        
        return {
            "chart1": {
                "title": "Расходы по категориям за текущий месяц",
                "description": f"Общая сумма расходов: {total_expenses:,.0f} ₽".replace(',', ' '),
                "data": expense_by_cat
            },
            "chart2": {
                "title": "Прогресс по финансовым целям",
                "description": f"Всего целей: {len(goals_data)}",
                "data": goals_data
            },
            "chart3": {
                "title": "Динамика чистого капитала за последние 12 недель",
                "description": f"Текущий чистый капитал: {weeks_data[-1]['net_capital']:,.0f} ₽".replace(',', ' ') if weeks_data else "Нет данных",
                "data": weeks_data
            }
        }


# Ценность 6: удаление всех данных (мои данные под контролем)
@app.delete("/api/me")
async def delete_my_account(user_id: int = Depends(get_user_id)):
    """Удалить все данные пользователя и аккаунт. Необратимо."""
    db = await get_db()
    async with db.acquire() as conn:
        try:
            await conn.execute("DELETE FROM budgets WHERE user_id = $1", user_id)
        except asyncpg.UndefinedTableError:
            pass
        await conn.execute("DELETE FROM ai_cache WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM ai_context WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM goals WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM transactions WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM asset_values WHERE asset_id IN (SELECT id FROM assets WHERE user_id = $1)", user_id)
        await conn.execute("DELETE FROM assets WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM liability_values WHERE liability_id IN (SELECT id FROM liabilities WHERE user_id = $1)", user_id)
        await conn.execute("DELETE FROM liabilities WHERE user_id = $1", user_id)
        await conn.execute("DELETE FROM users WHERE id = $1", user_id)
    return {"status": "ok", "message": "Аккаунт и все данные удалены."}


# Консультация - история по сессиям (1 сессия = 1 день = основная консультация + уточнения)
@app.get("/api/consultation/history")
async def get_consultation_history(user_id: int = Depends(get_user_id)):
    """Получить историю консультаций, сгруппированную по сессиям (дням)."""
    db = await get_db()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT role, content, created_at
            FROM ai_context
            WHERE user_id=$1
              AND (
                (role='assistant' AND content LIKE 'CONSULTATION:%')
                OR (role='user' AND content LIKE 'FOLLOW_UP:%')
              )
            ORDER BY created_at ASC
            LIMIT 500
            """,
            user_id
        )
    # Группируем по дате (день): в каждой сессии первое assistant CONSULTATION — основная консультация, пары user FOLLOW_UP + assistant CONSULTATION — уточнения
    from collections import defaultdict
    sessions_by_date = defaultdict(lambda: {"main": None, "follow_ups": []})
    for r in rows:
        dt = r["created_at"]
        day = dt.date() if hasattr(dt, "date") else dt
        if r["role"] == "assistant" and r["content"].startswith("CONSULTATION:"):
            text = r["content"].replace("CONSULTATION: ", "", 1)
            if sessions_by_date[day]["main"] is None:
                sessions_by_date[day]["main"] = {"content": text, "date": r["created_at"].isoformat()}
            else:
                # ответ на уточняющий вопрос
                fu = sessions_by_date[day]["follow_ups"]
                if fu and fu[-1].get("answer") is None:
                    fu[-1]["answer"] = text
                else:
                    fu.append({"question": None, "answer": text})
        elif r["role"] == "user" and r["content"].startswith("FOLLOW_UP:"):
            q = r["content"].replace("FOLLOW_UP: ", "", 1)
            sessions_by_date[day]["follow_ups"].append({"question": q, "answer": None})
    # Преобразуем в список сессий (от новых к старым)
    result = []
    for day in sorted(sessions_by_date.keys(), reverse=True)[:20]:
        s = sessions_by_date[day]
        if s["main"] is None:
            continue
        session = {
            "date": day.isoformat() if hasattr(day, "isoformat") else str(day),
            "content": s["main"]["content"],
            "follow_ups": [f for f in s["follow_ups"] if f.get("question") or f.get("answer")],
        }
        result.append(session)
    return result


# Консультация - проверка лимита по сессиям (1 сессия = 1 день; бесплатно 1/мес, по подписке 5/мес)
async def check_consultation_limit(user_id: int) -> tuple[bool, int, int, bool]:
    """
    Returns:
        tuple[bool, int, int, bool]: (can_request_main, sessions_used, limit, can_followup_today)
    """
    db = await get_db()
    now = datetime.now()
    since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    is_premium = await check_premium(user_id)
    limit = 5 if is_premium else 1

    async with db.acquire() as conn:
        # Количество сессий (уникальных дней) в этом месяце с хотя бы одной консультацией
        rows = await conn.fetch(
            """
            SELECT DISTINCT DATE(created_at) as d
            FROM ai_context
            WHERE user_id=$1 AND role='assistant' AND content LIKE 'CONSULTATION:%'
              AND created_at >= $2
            """,
            user_id, since
        )
        sessions_used = len(rows)
        # Последняя консультация — сегодня? (можно задать уточняющий вопрос)
        last_row = await conn.fetchrow(
            """
            SELECT DATE(created_at) as d
            FROM ai_context
            WHERE user_id=$1 AND role='assistant' AND content LIKE 'CONSULTATION:%'
            ORDER BY created_at DESC LIMIT 1
            """,
            user_id
        )
        last_date = last_row["d"] if last_row and last_row["d"] else None
        today = now.date()
        can_followup_today = last_date == today
    can_main = sessions_used < limit
    return can_main, sessions_used, limit, can_followup_today


# Консультация
@app.get("/api/consultation/limit")
async def get_consultation_limit(user_id: int = Depends(get_user_id)):
    """Лимит сессий (1 сессия = 1 день). Бесплатно 1/мес, по подписке 5/мес. И можно ли задать уточняющий вопрос сегодня."""
    can_main, sessions_used, limit, can_followup_today = await check_consultation_limit(user_id)
    return {
        "sessions_used": sessions_used,
        "limit": limit,
        "limit_reached": not can_main,
        "can_followup_today": can_followup_today,
    }


@app.get("/api/consultation")
async def get_consultation(user_id: int = Depends(get_user_id)):
    """Получить AI консультацию (основная консультация = 1 сессия в день; лимит 1/мес бесплатно, 5/мес по подписке)."""
    can_main, sessions_used, limit, _ = await check_consultation_limit(user_id)

    if not can_main:
        return {
            "consultation": None,
            "error": f"Лимит сессий исчерпан. Использовано: {sessions_used}/{limit} в этом месяце. Оформите подписку для 5 сессий в месяц.",
            "limit_reached": True,
            "sessions_used": sessions_used,
            "limit": limit,
        }

    try:
        logging.info(f"Consultation request for user_id={user_id} ({sessions_used + 1}/{limit})")
        consultation = await asyncio.wait_for(
            generate_consultation(user_id),
            timeout=60.0
        )
        logging.info("Consultation completed successfully")
        return {
            "consultation": consultation,
            "sessions_used": sessions_used + 1,
            "limit": limit,
            "limit_reached": sessions_used + 1 >= limit,
        }
    except asyncio.TimeoutError:
        logging.error("Consultation generation timeout (60s)")
        return {
            "consultation": (
                "⏱️ Генерация консультации заняла слишком много времени.\n\n"
                "Попробуйте позже."
            ),
            "sessions_used": sessions_used,
            "limit": limit,
            "limit_reached": False,
        }
    except Exception as e:
        logging.error(f"Error in consultation endpoint: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return {
            "consultation": (
                "❌ Произошла ошибка при генерации консультации.\n\n"
                "Попробуйте позже.\n\n"
                f"Ошибка: {str(e)[:100]}"
            ),
            "sessions_used": sessions_used,
            "limit": limit,
            "limit_reached": False,
        }


# Консультация — ввод целей через сообщение (AI извлекает цели и сохраняет в goals)
async def _extract_goals_from_message(user_message: str) -> list[dict]:
    """Вызвать GigaChat для извлечения финансовых целей из текста. Возвращает список { title, target, description }."""
    prompt = (
        "Пользователь написал сообщение о своих финансовых целях. Извлеки из текста цели.\n\n"
        "Цель — это намерение с суммой и/или сроком, например: накопить 1 000 000 за 2 года, "
        "пассивный доход 50 000 в месяц, погасить долг 200 000.\n\n"
        "ВАЖНО: Если пользователь упоминает ПАССИВНЫЙ ДОХОД (например, 'хочу получать 30 000 в месяц', "
        "'пассивный доход 50 тыс.', 'в 40 лет получать 100 000'), рассчитай необходимый капитал:\n"
        "- Для пассивного дохода используй правило 4% годовых: капитал = месячный_доход * 12 / 0.04\n"
        "- Например: пассивный доход 30 000 ₽/мес = капитал 9 000 000 ₽ (30 000 * 12 / 0.04)\n"
        "- В title укажи цель с упоминанием пассивного дохода, в target — рассчитанный капитал\n\n"
        "Ответь ТОЛЬКО валидным JSON-массивом объектов без комментариев:\n"
        '[{"title":"краткое название цели","target":число_в_рублях,"description":"описание или срок"}]'
        "\nЕсли целей нет — верни []. target — целевая сумма в рублях (число)."
    )
    try:
        messages = [
            {"role": "system", "content": "Ты извлекаешь финансовые цели из текста. Отвечай только JSON-массивом."},
            {"role": "user", "content": prompt + "\n\nСообщение пользователя:\n" + (user_message or "")[:2000]}
        ]
        answer = await gigachat_request(messages)
        answer = (answer or "").strip()
        import re
        json_match = re.search(r"\[[\s\S]*\]", answer)
        if not json_match:
            return []
        data = json.loads(json_match.group())
        result = []
        for item in data:
            if not isinstance(item, dict):
                continue
            try:
                title = str(item.get("title") or "Цель").strip() or "Цель"
                target = float(item.get("target", 0))
                if target <= 0:
                    continue
                description = (item.get("description") or "").strip() or None
                result.append({"title": title, "target": target, "description": description})
            except (TypeError, ValueError):
                continue
        return result
    except (json.JSONDecodeError, Exception):
        return []


@app.post("/api/consultation/message")
async def consultation_message(
    body: ConsultationMessageRequest,
    user_id: int = Depends(get_user_id)
):
    """Отправить сообщение в консультацию: AI извлекает цели и сохраняет в goals."""
    message = (body.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message is required")
    await save_message(user_id, "user", message)
    goals_added = []
    try:
        extracted = await _extract_goals_from_message(message)
        db = await get_db()
        async with db.acquire() as conn:
            for g in extracted:
                await conn.execute(
                    "INSERT INTO goals (user_id, target, current, title, description, created_at) VALUES ($1, $2, 0, $3, $4, NOW())",
                    user_id, g["target"], g["title"], g.get("description")
                )
                goals_added.append({"title": g["title"], "target": g["target"]})
    except Exception as e:
        logging.exception("consultation_message extract goals")
    reply = "Сообщение принято."
    if goals_added:
        reply = f"Цели добавлены: {', '.join(g['title'] + ' — ' + str(int(g['target'])) + ' ₽' for g in goals_added)}."
    return {"goals_added": goals_added, "reply": reply}


@app.post("/api/consultation/follow-up")
async def consultation_follow_up(
    body: ConsultationMessageRequest,
    user_id: int = Depends(get_user_id)
):
    """Задать уточняющий вопрос по последней консультации. Доступно только в тот же день (1 сессия = консультация + уточнения). Не тратит лимит сессий."""
    _, sessions_used, limit, can_followup_today = await check_consultation_limit(user_id)
    if not can_followup_today:
        raise HTTPException(
            status_code=400,
            detail="Уточняющий вопрос можно задать только в день получения консультации. Получите новую консультацию сегодня или завтра."
        )
    message = (body.message or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message is required")

    db = await get_db()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT content FROM ai_context
            WHERE user_id=$1 AND role='assistant' AND content LIKE 'CONSULTATION:%'
            ORDER BY created_at DESC LIMIT 1
            """,
            user_id
        )
    if not row:
        raise HTTPException(
            status_code=400,
            detail="Сначала получите консультацию (кнопка «Получить консультацию»)."
        )
    last_consultation = row["content"].replace("CONSULTATION: ", "", 1)

    system_prompt = (
        "Ты — персональный финансовый консультант. Пользователь задаёт уточняющий вопрос по своей последней консультации. "
        "Ответь кратко, по делу и дружелюбно на русском. Используй конкретные суммы и рекомендации. "
        "Формат чисел: с пробелами (200 000 ₽), без научной нотации и точек как разделителей."
    )
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": f"Последняя консультация:\n\n{last_consultation}\n\nВопрос пользователя: {message}"}
    ]
    try:
        answer = await asyncio.wait_for(gigachat_request(messages), timeout=40.0)
        if not answer or not answer.strip():
            answer = "Не удалось сформировать ответ. Попробуйте переформулировать вопрос."
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="Превышено время ожидания ответа.")
    except Exception as e:
        logging.exception("consultation_follow_up")
        raise HTTPException(status_code=500, detail="Ошибка при генерации ответа.")

    await save_message(user_id, "user", f"FOLLOW_UP: {message}")
    await save_message(user_id, "assistant", f"CONSULTATION: {answer}")
    return {
        "reply": answer,
        "sessions_used": sessions_used,
        "limit": limit,
        "limit_reached": sessions_used >= limit,
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
