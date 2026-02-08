# -*- coding: utf-8 -*-
"""Читает структуру и формулы из Финансовый калькулятор аренда или покупка.xlsx"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not found")
    sys.exit(1)

# Файл на рабочем столе пользователя
DESKTOP = os.path.expanduser("~") + os.path.sep + "OneDrive" + os.path.sep + "Desktop"
path = os.path.join(DESKTOP, "Финансовый калькулятор аренда или покупка.xlsx")
if not os.path.isfile(path):
    path = os.path.join(os.path.expanduser("~"), "Desktop", "Финансовый калькулятор аренда или покупка.xlsx")
if not os.path.isfile(path):
    print("File not found:", path)
    sys.exit(1)

wb = openpyxl.load_workbook(path, data_only=False)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=== Sheet:", sheet_name, "===")
    for row in range(1, min(ws.max_row + 1, 60)):
        for col in range(1, min(ws.max_column + 1, 20)):
            c = ws.cell(row, col)
            v = c.value
            if v is not None and str(v).strip():
                is_formula = c.data_type == "f"
                print(f"  {row},{col} ({c.column_letter}{row}): {repr(v)[:80]} formula={is_formula}")
wb.close()
print("Done.")
